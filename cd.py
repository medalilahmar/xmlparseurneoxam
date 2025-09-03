import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox, QFileDialog, QFrame, QRadioButton, QButtonGroup, QDialog, QTableWidget, QTableWidgetItem ,QProgressDialog
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import os
import re
import hashlib
import bcrypt
import uuid
from PyQt5.QtWidgets import QWidget, QPushButton, QHBoxLayout, QVBoxLayout, QSizePolicy
from PyQt5.QtCore import Qt, QSize
import sqlite3
import xml.etree.ElementTree as ET
import csv
import random as rd
import qrcode
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from subprocess import Popen
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import urllib.parse
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from wordcloud import WordCloud
from collections import Counter
import numpy as np
import mplcursors
import os
import csv
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QCheckBox, QMessageBox, QTableWidget, QInputDialog
from PyQt5.QtCore import Qt, QThread, pyqtSignal

def adapt_datetime(dt):
    return dt.isoformat()

sqlite3.register_adapter(datetime, adapt_datetime)

class OptionDatabaseManager:
    """Enhanced database manager for XML file storage and option tracking"""

    def __init__(self, db_path="option_search.db"):
        self.db_path = db_path
        self.init_database()

    def init_database(self):
        """Initialize database with enhanced schema and handle migrations"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS xml_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT NOT NULL,
                    filepath TEXT NOT NULL UNIQUE,  -- Modified to ensure filepath uniqueness
                    processed_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    file_hash TEXT,
                    file_size INTEGER,
                    last_modified TIMESTAMP
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS options (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER,
                    option_name TEXT NOT NULL,
                    procedure_name TEXT,
                    is_covered BOOLEAN DEFAULT 0,
                    coverage_details TEXT,
                    treatment_code TEXT,
                    UNIQUE(file_id, option_name, treatment_code),
                    FOREIGN KEY (file_id) REFERENCES xml_files (id) ON DELETE CASCADE
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS variables (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    option_id INTEGER,
                    variable_name TEXT NOT NULL,
                    variable_value TEXT,
                    variable_type TEXT,
                    is_required BOOLEAN DEFAULT 0,
                    description TEXT,
                    FOREIGN KEY (option_id) REFERENCES options (id) ON DELETE CASCADE
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS treatment_codes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT UNIQUE NOT NULL,
                    description TEXT,
                    is_active BOOLEAN DEFAULT 1,
                    category TEXT
                )
            """)

            conn.commit()
        except sqlite3.Error as e:
            print(f"Database initialization error: {e}")
            raise
        finally:
            conn.close()

    def add_xml_file(self, filename, filepath, file_hash):
        """Add or update XML file in database with enhanced metadata"""
        try:
            file_size = os.path.getsize(filepath)
            last_modified = datetime.fromtimestamp(os.path.getmtime(filepath))

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Check if a file with the same filepath already exists
            cursor.execute("SELECT id FROM xml_files WHERE filepath = ?", (filepath,))
            existing_file = cursor.fetchone()

            if existing_file:
                # Update existing record
                file_id = existing_file[0]
                cursor.execute("""
                    UPDATE xml_files
                    SET filename = ?, file_hash = ?, file_size = ?, last_modified = ?, processed_date = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (filename, file_hash, file_size, last_modified, file_id))
                print(f"[DEBUG] Updated existing file: filepath={filepath}, file_id={file_id}")
            else:
                # Insert new record
                cursor.execute("""
                    INSERT INTO xml_files (filename, filepath, file_hash, file_size, last_modified)
                    VALUES (?, ?, ?, ?, ?)
                """, (filename, filepath, file_hash, file_size, last_modified))
                file_id = cursor.lastrowid
                print(f"[DEBUG] Inserted new file: filepath={filepath}, file_id={file_id}")

            conn.commit()
            return file_id
        except sqlite3.Error as e:
            print(f"Error adding/updating XML file to database: {e}")
            raise
        finally:
            conn.close()

    def add_option(self, file_id, option_name, procedure_name, is_covered, treatment_code, coverage_details=""):
        """Add option to database, avoiding duplicates"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                INSERT OR IGNORE INTO options (file_id, option_name, procedure_name, is_covered, treatment_code, coverage_details)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (file_id, option_name, procedure_name, is_covered, treatment_code, coverage_details))

            cursor.execute("""
                SELECT id FROM options WHERE file_id = ? AND option_name = ? AND treatment_code = ?
            """, (file_id, option_name, treatment_code))
            result = cursor.fetchone()
            option_id = result[0] if result else None

            conn.commit()
            return option_id
        except sqlite3.Error as e:
            print(f"Error adding option to database: {e}")
            raise
        finally:
            conn.close()

    def add_variable(self, option_id, variable_name, variable_value, variable_type, is_required, description=""):
        """Add variable to database"""
        if not option_id:
            return
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO variables (option_id, variable_name, variable_value, variable_type, is_required, description)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (option_id, variable_name, variable_value, variable_type, is_required, description))

            conn.commit()
        except sqlite3.Error as e:
            print(f"Error adding variable to database: {e}")
            raise
        finally:
            conn.close()

    def add_treatment_code(self, code, description="", category=""):
        """Add treatment code to database"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                INSERT OR IGNORE INTO treatment_codes (code, description, category, is_active)
                VALUES (?, ?, ?, ?)
            """, (code, description, category, 1))

            conn.commit()
        except sqlite3.Error as e:
            print(f"Error adding treatment code to database: {e}")
            raise
        finally:
            conn.close()

    def search_options(self, option_name, treatment_code=None):
        """Enhanced search for options with filtering"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            query = """
                SELECT o.option_name, o.procedure_name, o.is_covered, o.coverage_details,
                       o.treatment_code, xf.filename, xf.filepath, xf.id
                FROM options o
                JOIN xml_files xf ON o.file_id = xf.id
                WHERE o.option_name LIKE ?
            """
            params = [f"%{option_name}%"]

            if treatment_code and treatment_code != "all":
                query += " AND o.treatment_code = ?"
                params.append(treatment_code)

            cursor.execute(query, params)
            results = cursor.fetchall()
            return results
        except sqlite3.Error as e:
            print(f"Error searching options: {e}")
            raise
        finally:
            conn.close()

    def get_treatment_codes(self):
        """Get all treatment codes"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT code FROM treatment_codes WHERE is_active = 1")
            codes = [row[0] for row in cursor.fetchall()]
            return codes
        except sqlite3.Error as e:
            print(f"Error retrieving treatment codes: {e}")
            raise
        finally:
            conn.close()

    def delete_xml_file(self, file_id):
        """Delete an XML file and its associated options and variables"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("DELETE FROM xml_files WHERE id = ?", (file_id,))
            conn.commit()

            return cursor.rowcount > 0  # Returns True if a row was deleted
        except sqlite3.Error as e:
            print(f"Error deleting XML file from database: {e}")
            raise
        finally:
            conn.close()



class DatabaseManager:
    """Gère la base de données SQLite pour les utilisateurs, sessions et tentatives de connexion."""
    
    def __init__(self, db_path="users.db"):
        """Initialise le gestionnaire de base de données."""
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialise la base de données avec les tables nécessaires."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT UNIQUE NOT NULL,
                        email TEXT UNIQUE,
                        password TEXT NOT NULL,
                        role TEXT DEFAULT 'user',
                        active INTEGER DEFAULT 0,
                        locked INTEGER DEFAULT 0,
                        failed_attempts INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS logins (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        username TEXT,
                        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        status TEXT,
                        ip_address TEXT,
                        FOREIGN KEY (user_id) REFERENCES users (id)
                    )
                ''')
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS sessions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        session_token TEXT UNIQUE,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        expires_at TIMESTAMP,
                        active INTEGER DEFAULT 1,
                        FOREIGN KEY (user_id) REFERENCES users (id)
                    )
                ''')
                cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', ('admin',))
                if cursor.fetchone()[0] == 0:
                    admin_password = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt())
                    cursor.execute('''
                        INSERT INTO users (username, email, password, role, active)
                        VALUES (?, ?, ?, ?, ?)
                    ''', ("admin", "admin@example.com", admin_password, "admin", 1))
                conn.commit()
                print(f"[DEBUG] Database initialized at {self.db_path}")
        except sqlite3.Error as e:
            print(f"[ERROR] Failed to initialize database: {e}")
            raise
    
    def validate_email(self, email):
        """Valide le format de l'adresse email."""
        if not email:
            return False
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    def email_exists(self, email):
        """Vérifie si une adresse email existe déjà dans la base."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT 1 FROM users WHERE email = ?', (email,))
                return cursor.fetchone() is not None
        except sqlite3.Error as e:
            print(f"[ERROR] Error checking email existence: {e}")
            return False
    
    def create_user(self, username, password, email, role="user"):
        """Crée un nouvel utilisateur avec validation renforcée."""
        print(f"[DEBUG] Attempting to create user: username={username}, email={email}, role={role}")
        if not (3 <= len(username) <= 20 and username.isalnum()):
            print(f"[ERROR] Invalid username: {username}")
            return False, "Le nom d'utilisateur doit avoir entre 3 et 20 caractères alphanumériques."
        if len(password) < 6:
            print(f"[ERROR] Password too short for username: {username}")
            return False, "Le mot de passe doit contenir au moins 6 caractères."
        if not self.validate_email(email):
            print(f"[ERROR] Invalid email: {email}")
            return False, "L'adresse email n'est pas valide."
        if self.email_exists(email):
            print(f"[ERROR] Email already exists: {email}")
            return False, "L'adresse email est déjà utilisée."
        try:
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO users (username, email, password, role, active)
                    VALUES (?, ?, ?, ?, ?)
                ''', (username, email, hashed_password, role, 0))
                conn.commit()
                print(f"[INFO] User created: username={username}, email={email}")
                return True, "Utilisateur créé avec succès. En attente de validation par l'admin."
        except sqlite3.IntegrityError as e:
            print(f"[ERROR] Database integrity error: {e}")
            return False, "Nom d'utilisateur ou email déjà existant."
        except sqlite3.Error as e:
            print(f"[ERROR] Database error during user creation: {e}")
            return False, f"Erreur lors de la création: {str(e)}"
    
    def authenticate_user(self, username, password):
        """Authentifie un utilisateur."""
        print(f"[DEBUG] Authenticating user: username={username}")
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT id, username, email, password, role, active, locked, failed_attempts
                    FROM users WHERE username = ?
                ''', (username,))
                user = cursor.fetchone()
                if not user:
                    self.log_login_attempt(None, username, "user_not_found")
                    print(f"[ERROR] User not found: {username}")
                    return False, "Nom d'utilisateur inexistant."
                user_id, db_username, email, db_password, role, active, locked, failed_attempts = user
                if locked:
                    self.log_login_attempt(user_id, username, "account_locked")
                    print(f"[ERROR] Account locked: username={username}")
                    return False, "Compte verrouillé. Contactez l'administrateur."
                if not bcrypt.checkpw(password.encode('utf-8'), db_password):
                    failed_attempts += 1
                    if failed_attempts >= 3:
                        cursor.execute('UPDATE users SET locked = 1, failed_attempts = ? WHERE id = ?', 
                                     (failed_attempts, user_id))
                        conn.commit()
                        self.log_login_attempt(user_id, username, "account_locked_max_attempts")
                        print(f"[ERROR] Account locked after max attempts: username={username}")
                        return False, "Compte verrouillé après 3 tentatives échouées."
                    else:
                        cursor.execute('UPDATE users SET failed_attempts = ? WHERE id = ?', 
                                     (failed_attempts, user_id))
                        conn.commit()
                        self.log_login_attempt(user_id, username, "failed_password")
                        print(f"[ERROR] Incorrect password for username={username}, attempts={failed_attempts}")
                        return False, f"Mot de passe incorrect. {3-failed_attempts} tentatives restantes."
                if not active:
                    self.log_login_attempt(user_id, username, "account_not_active")
                    print(f"[ERROR] Account not active: username={username}")
                    return False, "Compte en attente de validation par l'administrateur."
                cursor.execute('UPDATE users SET failed_attempts = 0 WHERE id = ?', (user_id,))
                conn.commit()
                session_token = str(uuid.uuid4())
                expires_at = datetime.now() + timedelta(hours=8)
                cursor.execute('''
                    INSERT INTO sessions (user_id, session_token, expires_at)
                    VALUES (?, ?, ?)
                ''', (user_id, session_token, expires_at))
                conn.commit()
                self.log_login_attempt(user_id, username, "success")
                print(f"[INFO] User authenticated: username={username}, role={role}")
                return True, {
                    "user_id": user_id,
                    "username": username,
                    "email": email,
                    "role": role,
                    "session_token": session_token
                }
        except sqlite3.Error as e:
            print(f"[ERROR] Database error during authentication: {e}")
            return False, f"Erreur d'authentification: {str(e)}"
    
    def log_login_attempt(self, user_id, username, status, ip_address=None):
        """Enregistre une tentative de connexion."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO logins (user_id, username, status, ip_address)
                    VALUES (?, ?, ?, ?)
                ''', (user_id, username, status, ip_address))
                conn.commit()
                print(f"[DEBUG] Logged login attempt: username={username}, status={status}, ip={ip_address}")
        except sqlite3.Error as e:
            print(f"[ERROR] Error logging login attempt: {e}")
    
    def get_all_users(self):
        """Récupère tous les utilisateurs pour l'interface admin."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT id, username, email, role, active, locked, failed_attempts, created_at
                    FROM users ORDER BY created_at DESC
                ''')
                users = cursor.fetchall()
                print(f"[DEBUG] Retrieved {len(users)} users from database")
                return users
        except sqlite3.Error as e:
            print(f"[ERROR] Error fetching users: {e}")
            return []
    
    def get_user_by_id(self, user_id):
        """Récupère les détails d'un utilisateur par ID."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT id, username, email, role, active, locked, failed_attempts, created_at
                    FROM users WHERE id = ?
                ''', (user_id,))
                user = cursor.fetchone()
                if user:
                    print(f"[DEBUG] Retrieved user by ID: id={user_id}")
                else:
                    print(f"[ERROR] User not found: id={user_id}")
                return user
        except sqlite3.Error as e:
            print(f"[ERROR] Error fetching user by ID: {e}")
            return None
    
    def update_user_status(self, user_id, active=None, locked=None, role=None):
        """Met à jour le statut d'un utilisateur."""
        print(f"[DEBUG] Updating user status: id={user_id}, active={active}, locked={locked}, role={role}")
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                updates = []
                params = []
                if active is not None:
                    updates.append("active = ?")
                    params.append(active)
                if locked is not None:
                    updates.append("locked = ?")
                    params.append(locked)
                    if not locked:
                        updates.append("failed_attempts = 0")
                if role is not None:
                    updates.append("role = ?")
                    params.append(role)
                if updates:
                    query = f"UPDATE users SET {', '.join(updates)} WHERE id = ?"
                    params.append(user_id)
                    cursor.execute(query, params)
                    conn.commit()
                    print(f"[INFO] User status updated: id={user_id}")
                    return True
                else:
                    print(f"[ERROR] No updates provided for user id={user_id}")
                    return False
        except sqlite3.Error as e:
            print(f"[ERROR] Error updating user status: {e}")
            return False
    
    def reset_password(self, user_id, new_password):
        """Réinitialise le mot de passe d'un utilisateur."""
        print(f"[DEBUG] Resetting password for user id={user_id}")
        try:
            if len(new_password) < 6:
                print(f"[ERROR] New password too short for user id={user_id}")
                return False, "Le mot de passe doit contenir au moins 6 caractères."
            hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE users SET password = ?, failed_attempts = 0, locked = 0
                    WHERE id = ?
                ''', (hashed_password, user_id))
                conn.commit()
                print(f"[INFO] Password reset for user id={user_id}")
                return True, "Mot de passe réinitialisé avec succès."
        except sqlite3.Error as e:
            print(f"[ERROR] Error resetting password: {e}")
            return False, f"Erreur lors de la réinitialisation: {str(e)}"
    
    def delete_user(self, user_id):
        """Supprime un utilisateur et ses données associées (sessions, logins)."""
        print(f"[DEBUG] Attempting to delete user: id={user_id}")
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                # Vérifier si l'utilisateur existe
                cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
                user = cursor.fetchone()
                if not user:
                    print(f"[ERROR] User not found: id={user_id}")
                    return False, "Utilisateur non trouvé."
                
                # Vérifier si l'utilisateur est l'admin
                if user[0] == 'admin':
                    print(f"[ERROR] Cannot delete admin user: id={user_id}")
                    return False, "Impossible de supprimer l'utilisateur admin."
                
                # Supprimer les sessions associées
                cursor.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
                print(f"[DEBUG] Deleted sessions for user id={user_id}")
                
                # Supprimer les tentatives de connexion associées
                cursor.execute('DELETE FROM logins WHERE user_id = ?', (user_id,))
                print(f"[DEBUG] Deleted logins for user id={user_id}")
                
                # Supprimer l'utilisateur
                cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
                conn.commit()
                print(f"[INFO] User deleted successfully: id={user_id}")
                return True, "Utilisateur supprimé avec succès."
        except sqlite3.Error as e:
            print(f"[ERROR] Error deleting user: {e}")
            return False, f"Erreur lors de la suppression de l'utilisateur: {str(e)}"



class RegisterWindow(QWidget):
    """Fenêtre pour l'inscription de nouveaux utilisateurs avec validation et envoi d'email."""

    def __init__(self):
        """Initialise la fenêtre d'inscription."""
        super().__init__()
        self.db_manager = DatabaseManager()
        self.email_manager = EmailManager()
        self.configure_email_settings()
        self.initUI()
        print("[DEBUG] RegisterWindow initialized")

    def configure_email_settings(self):
        """Configure les paramètres SMTP pour l'envoi d'emails."""
        self.email_manager.configure_smtp(
            smtp_server="smtp.gmail.com",
            smtp_port=587,
            sender_email="medalilahmar00@gmail.com",  # ⚠️ Remplacer par votre email
            sender_password="tlwu nhmm japb dgnu",  # ⚠️ Remplacer par votre mot de passe d'application
            app_name="PyQt5"
        )
        logo_path = "images/pdflogo.png"
        self.logo_path = logo_path if os.path.exists(logo_path) else None
        print(f"[DEBUG] RegisterWindow Logo path: {self.logo_path}")

    def initUI(self):
        """Initialise l'interface utilisateur de la fenêtre d'inscription."""
        self.setWindowTitle('Inscription')
        self.setWindowIcon(QIcon("images/iconapp.ico"))
        self.setFixedSize(600, 700)
        self.setStyleSheet("""
            QWidget {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #000000,
                    stop: 0.4 #1a1a2e,
                    stop: 1 #0f2f2f
                );
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                color: #FFFFFF;
                font-size: 16px;
                font-weight: bold;
                background: transparent;
            }
            QLineEdit {
                background-color: #1C2526;
                color: #FFFFFF;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 12px;
                padding: 12px;
                font-size: 14px;
                selection-background-color: #12e0d6;
            }
            QLineEdit:focus {
                border: 2px solid #0fb8b0;
                outline: none;
                background-color: #1C2526;
            }
            QPushButton {
                background-color: #12e0d6;
                color: #000000;
                border: 2px solid #0fb8b0;
                border-radius: 12px;
                padding: 12px 20px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0fb8b0;
                border-color: #0d8f88;
            }
            QPushButton:pressed {
                background-color: #0d8f88;
            }
            QPushButton:disabled {
                background-color: #666666;
                color: #999999;
                border-color: #555555;
            }
            QCheckBox {
                color: #FFFFFF;
                font-size: 14px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                background-color: #1C2526;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 4px;
            }
            QCheckBox::indicator:checked {
                background-color: #12e0d6;
                border-color: #0fb8b0;
            }
            QHBoxLayout {
                spacing: 15px;
            }
        """)

        layout = QVBoxLayout()
        
        # Titre
        title = QLabel('Créer un nouveau compte')
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold; margin-bottom: 20px;")
        layout.addWidget(title)
        
        # Champ Nom d'utilisateur
        self.username_label = QLabel('Nom d\'utilisateur:')
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText('3-20 caractères alphanumériques')
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)
        
        # Champ Email
        self.email_label = QLabel('Adresse email:')
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText('votre-email@exemple.com')
        layout.addWidget(self.email_label)
        layout.addWidget(self.email_input)
        
        # Champ Mot de passe
        self.password_label = QLabel('Mot de passe:')
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText('Au moins 6 caractères')
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        
        # Champ Confirmer le mot de passe
        self.confirm_password_label = QLabel('Confirmer le mot de passe:')
        self.confirm_password_input = QLineEdit()
        self.confirm_password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.confirm_password_label)
        layout.addWidget(self.confirm_password_input)
        
        # Checkbox pour l'envoi d'email
        self.email_notification_checkbox = QCheckBox('📧 Recevoir un email de confirmation')
        self.email_notification_checkbox.setChecked(True)
        self.email_notification_checkbox.setStyleSheet("margin: 10px 0;")
        layout.addWidget(self.email_notification_checkbox)
        
        # Boutons
        button_layout = QHBoxLayout()
        self.register_button = QPushButton('S\'inscrire')
        self.register_button.clicked.connect(self.register_user)
        button_layout.addWidget(self.register_button)
        
        self.back_button = QPushButton('Retour')
        self.back_button.clicked.connect(self.back_to_login)
        button_layout.addWidget(self.back_button)
        
        self.cancel_button = QPushButton('Annuler')
        self.cancel_button.clicked.connect(self.close)
        button_layout.addWidget(self.cancel_button)
        
        layout.addLayout(button_layout)
        
        # Étiquette de statut
        self.status_label = QLabel('')
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("""
            color: #12e0d6;
            font-size: 12px;
            margin-top: 10px;
            padding: 5px;
            background: rgba(18, 224, 214, 0.1);
            border-radius: 5px;
        """)
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)

    def back_to_login(self):
        """Revient à la fenêtre de connexion."""
        print("[DEBUG] Navigating back to LoginWindow")
        self.close()
        self.login_window = LoginWindow()
        self.login_window.show()

    def register_user(self):
        """Enregistre un nouvel utilisateur et envoie un email de confirmation si activé."""
        username = self.username_input.text().strip()
        email = self.email_input.text().strip()
        password = self.password_input.text()
        confirm_password = self.confirm_password_input.text()
        
        print(f"[DEBUG] Registering user: username={username}, email={email}")
        
        # Validation des champs
        if not username or not email or not password:
            QMessageBox.warning(self, 'Erreur', 'Veuillez remplir tous les champs.')
            print("[ERROR] Missing fields")
            return
        
        if not self.db_manager.validate_email(email):
            QMessageBox.warning(self, 'Erreur', 'Veuillez entrer une adresse email valide.')
            print(f"[ERROR] Invalid email: {email}")
            return
        
        if len(password) < 6:
            QMessageBox.warning(self, 'Erreur', 'Le mot de passe doit contenir au moins 6 caractères.')
            print("[ERROR] Password too short")
            return
        
        if password != confirm_password:
            QMessageBox.warning(self, 'Erreur', 'Les mots de passe ne correspondent pas.')
            print("[ERROR] Passwords do not match")
            return
        
        # Désactiver le bouton pendant l'inscription
        self.register_button.setEnabled(False)
        self.register_button.setText('⏳ Inscription en cours...')
        self.status_label.setText('💾 Création du compte...')
        
        # Créer l'utilisateur dans la base de données
        success, message = self.db_manager.create_user(username, password, email)
        print(f"[DEBUG] Database create_user result: success={success}, message={message}")
        
        if success:
            if self.email_notification_checkbox.isChecked():
                self.status_label.setText('📧 Envoi de l\'email de confirmation...')
                self.send_confirmation_email(email, username)
            else:
                QMessageBox.information(self, 'Succès', 'Compte créé avec succès! En attente de validation.')
                print("[INFO] User created, no email sent (checkbox unchecked)")
                self.reset_form()
        else:
            QMessageBox.warning(self, 'Erreur', message)
            self.register_button.setEnabled(True)
            self.register_button.setText('S\'inscrire')
            self.status_label.setText('')
            print(f"[ERROR] User creation failed: {message}")

    def send_confirmation_email(self, recipient_email, username):
        """Envoie un email de confirmation dans un thread séparé."""
        print(f"[DEBUG] Initiating confirmation email to {recipient_email}, logo_path={self.logo_path}")
        logo_path = None
        if self.logo_path and os.path.exists(self.logo_path):
            logo_path = self.logo_path
        self.email_thread = EmailSendThread(
            self.email_manager,
            recipient_email,
            username,
            logo_path,
            email_type="confirmation"
        )
        self.email_thread.email_sent.connect(self.on_email_sent)
        self.email_thread.start()

    def on_email_sent(self, success, message):
        """Gère le résultat de l'envoi de l'email."""
        print(f"[DEBUG] Email send result: success={success}, message={message}")
        self.register_button.setEnabled(True)
        self.register_button.setText('S\'inscrire')
        
        if success:
            self.status_label.setText('✅ Email de confirmation envoyé!')
            QMessageBox.information(
                self,
                'Succès',
                f'🎉 Compte créé avec succès!\n\n📧 Un email de confirmation a été envoyé à:\n{self.email_input.text()}'
            )
        else:
            self.status_label.setText('⚠️ Compte créé, erreur email')
            QMessageBox.information(
                self,
                'Compte créé',
                f'✅ Votre compte a été créé avec succès!\n\n⚠️ Note: L\'email de confirmation n\'a pas pu être envoyé:\n{message}\n\nVérifiez la configuration email.'
            )
        
        self.reset_form()

    def reset_form(self):
        """Réinitialise les champs du formulaire."""
        print("[DEBUG] Resetting form")
        self.username_input.clear()
        self.email_input.clear()
        self.password_input.clear()
        self.confirm_password_input.clear()
        self.status_label.setText('')








import csv
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFrame, QLabel, QLineEdit, 
    QComboBox, QPushButton, QTableWidget, QTableWidgetItem, 
    QFileDialog, QMessageBox
)
from PyQt5.QtCore import Qt

class OptionSearchDialog(QDialog):
    """Enhanced option search dialog with modern design"""
    
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.parent = parent
        self.initUI()
    
    def initUI(self):
        """Initialize search dialog UI with modern design"""
        self.setWindowTitle("Recherche d'Options")
        self.setWindowIcon(QIcon("images/iconapp.ico"))
        self.setFixedSize(1200, 900)
        self.setStyleSheet("""
               QDialog {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #000000,
                    stop: 0.4 #1a1a2e,
                    stop: 1 #0f2f2f
                );
                font-family: 'Segoe UI', Arial, sans-serif;
                color: #ffffff;
                border: 2px solid rgba(18, 224, 214, 0.2);
                border-radius: 16px;
                backdrop-filter: blur(8px);
            }

            /* Labels */
            QLabel {
                color: #ffffff;
                font-size: 12px;
                font-weight: 500;
                background: transparent;
                margin: 2px 0;
            }

            QLabel[objectName="titleLabel"] {
                font-size: 18px;
                font-weight: 700;
                color: #12e0d6;
                text-align: center;
                margin: 10px 0;
                text-shadow: 0 0 10px rgba(18, 224, 214, 0.3);
            }

            QLabel[objectName="sectionLabel"] {
                font-size: 13px;
                font-weight: 600;
                color: #12e0d6;
                margin: 6px 0 4px 0;
            }

            /* Input fields */
            QLineEdit {
                background: rgba(28, 37, 38, 0.9);
                color: #ffffff;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 8px;
                padding: 8px 10px;
                font-size: 11px;
                selection-background-color: rgba(18, 224, 214, 0.4);
                backdrop-filter: blur(10px);
            }

            QLineEdit:focus {
                border: 2px solid #12e0d6;
                background: rgba(28, 37, 38, 0.95);
                box-shadow: 0 0 6px rgba(18, 224, 214, 0.3);
            }

            QLineEdit:hover {
                border: 2px solid rgba(18, 224, 214, 0.5);
            }

            QLineEdit::placeholder {
                color: #94a3b8;
                font-style: italic;
            }

            /* ComboBox */
            QComboBox {
                background: rgba(28, 37, 38, 0.9);
                color: #ffffff;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 8px;
                padding: 8px 10px;
                font-size: 11px;
                min-height: 32px;
                backdrop-filter: blur(10px);
            }

            QComboBox:hover {
                border: 2px solid rgba(18, 224, 214, 0.5);
            }

            QComboBox:focus {
                border: 2px solid #12e0d6;
                box-shadow: 0 0 6px rgba(18, 224, 214, 0.3);
            }

            QComboBox::drop-down {
                border: none;
                width: 30px;
            }

            QComboBox::down-arrow {
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #12e0d6;
                margin-right: 8px;
            }

            QComboBox:on::down-arrow {
                border-top: 6px solid #0fb8b0;
                transform: rotate(180deg);
            }

            QComboBox QAbstractItemView {
                background: rgba(28, 37, 38, 0.95);
                color: #ffffff;
                selection-background-color: rgba(18, 224, 214, 0.3);
                selection-color: #12e0d6;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 8px;
                padding: 4px;
                backdrop-filter: blur(10px);
            }

            /* Buttons */
            QPushButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                color: #000000;
                border: 2px solid #0fb8b0;
                border-radius: 8px;
                padding: 8px 16px;
                font-size: 11px;
                font-weight: 600;
                min-height: 32px;
                min-width: 100px;
                box-shadow: 0 3px 8px rgba(18, 224, 214, 0.2);
                transition: all 0.3s ease;
            }

            QPushButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #0fb8b0,
                    stop: 1 #0d8f88
                );
                transform: translateY(-2px);
                box-shadow: 0 5px 12px rgba(18, 224, 214, 0.3);
            }

            QPushButton:pressed {
                background: #0d8f88;
                transform: translateY(0);
                box-shadow: 0 2px 5px rgba(18, 224, 214, 0.2);
            }

            QPushButton[objectName="cancelButton"] {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #475569,
                    stop: 1 #374151
                );
                color: #ffffff;
                border: 2px solid #475569;
            }

            QPushButton[objectName="cancelButton"]:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #64748b,
                    stop: 1 #475569
                );
            }

            QPushButton[objectName="deleteButton"] {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #ef4444,
                    stop: 1 #dc2626
                );
                color: #ffffff;
                border: 2px solid #dc2626;
            }

            QPushButton[objectName="deleteButton"]:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #f87171,
                    stop: 1 #ef4444
                );
                box-shadow: 0 5px 12px rgba(239, 68, 68, 0.3);
            }

            /* Table */
            QTableWidget {
                background: rgba(28, 37, 38, 0.9);
                color: #e6e6e6;
                border: 1.5px solid rgba(18, 224, 214, 0.4);
                border-radius: 16px;
                gridline-color: rgba(71, 85, 105, 0.6);
                font-size: 17px;
                font-weight: 500;
                selection-background-color: rgba(18, 224, 214, 0.35);
                backdrop-filter: blur(12px);
            }

            QTableWidget::item {
                padding: 14px 16px;
                border-bottom: 1px solid rgba(71, 85, 105, 0.6);
                min-height: 60px;
            }

            QTableWidget::item:alternate {
                background: rgba(28, 37, 38, 0.3);
            }

            QTableWidget::item:selected {
                background: rgba(18, 224, 214, 0.25);
                color: #12e0d6;
                border-left: 4px solid #12e0d6;
            }

            QTableWidget::item:hover {
                background: rgba(18, 224, 214, 0.2);
            }

            QHeaderView::section {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                color: #000000;
                font-weight: 600;
                font-size: 16px;
                padding: 14px 16px;
                border: none;
                border-radius: 0;
                text-transform: uppercase;
            }

            QHeaderView::section:first {
                border-top-left-radius: 14px;
            }

            QHeaderView::section:last {
                border-top-right-radius: 14px;
            }

            /* Frame */
            QFrame[objectName="cardFrame"] {
                background: rgba(28, 37, 38, 0.4);
                border: 1px solid rgba(18, 224, 214, 0.15);
                border-radius: 12px;
                padding: 10px;
                margin: 4px;
                backdrop-filter: blur(10px);
                box-shadow: 0 4px 16px rgba(0, 0, 0, 0.2);
            }

            /* Scrollbars */
            QScrollBar:vertical {
                background: rgba(28, 37, 38, 0.3);
                width: 10px;
                border-radius: 5px;
            }

            QScrollBar::handle:vertical {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                border-radius: 5px;
                min-height: 25px;
                margin: 2px;
            }

            QScrollBar::handle:vertical:hover {
                background: #0fb8b0;
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0;
            }

            QScrollBar:horizontal {
                background: rgba(28, 37, 38, 0.3);
                height: 10px;
                border-radius: 5px;
            }

            QScrollBar::handle:horizontal {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                border-radius: 5px;
                min-width: 25px;
                margin: 2px;
            }

            QScrollBar::handle:horizontal:hover {
                background: #0fb8b0;
            }

            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                width: 0;
            }

            /* Tooltip */
            QToolTip {
                background: rgba(28, 37, 38, 0.95);
                color: #ffffff;
                border: 1px solid rgba(18, 224, 214, 0.3);
                border-radius: 6px;
                padding: 6px 10px;
                font-size: 11px;
                backdrop-filter: blur(10px);
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        # Header
        header_frame = QFrame()
        header_frame.setObjectName('cardFrame')
        header_layout = QHBoxLayout(header_frame)
        title_label = QLabel("🔍 Recherche d'Options")
        title_label.setObjectName('titleLabel')
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        layout.addWidget(header_frame)

        # Search input section
        search_card = QFrame()
        search_card.setObjectName('cardFrame')
        search_layout = QHBoxLayout(search_card)
        search_layout.setSpacing(12)
        search_label = QLabel("Rechercher une Option:")
        search_label.setObjectName('sectionLabel')
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Entrez le nom de l'option...")
        self.search_input.setFixedHeight(48)
        search_button = QPushButton("🔍 Rechercher")
        search_button.setFixedSize(160, 48)
        search_button.clicked.connect(self.search_options)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input, stretch=1)
        search_layout.addWidget(search_button)
        layout.addWidget(search_card)

        # Treatment code filter
        filter_card = QFrame()
        filter_card.setObjectName('cardFrame')
        filter_layout = QHBoxLayout(filter_card)
        filter_layout.setSpacing(12)
        treatment_label = QLabel("Filtrer par Code de Traitement:")
        treatment_label.setObjectName('sectionLabel')
        self.treatment_combo = QComboBox()
        self.treatment_combo.setFixedHeight(48)
        self.treatment_combo.addItem("Tous les Codes", "all")
        try:
            codes = self.db_manager.get_treatment_codes()
            for code in codes:
                self.treatment_combo.addItem(code, code)
        except Exception as e:
            self.log_message(f"Erreur lors du chargement des codes de traitement: {e}")
        filter_layout.addWidget(treatment_label)
        filter_layout.addWidget(self.treatment_combo)
        filter_layout.addStretch()
        layout.addWidget(filter_card)

        # Results table
        table_card = QFrame()
        table_card.setObjectName('cardFrame')
        table_layout = QVBoxLayout(table_card)
        table_title = QLabel("Résultats de la Recherche")
        table_title.setObjectName('sectionLabel')
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(4)
        self.results_table.setHorizontalHeaderLabels(["Option", "Fichier", "Chemin", "File ID"])
        self.results_table.setColumnHidden(3, True)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setShowGrid(True)
        self.results_table.setSelectionMode(QTableWidget.SingleSelection)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        table_layout.addWidget(table_title)
        table_layout.addWidget(self.results_table)
        layout.addWidget(table_card, stretch=1)

        # Buttons
        buttons_card = QFrame()
        buttons_card.setObjectName('cardFrame')
        buttons_layout = QHBoxLayout(buttons_card)
        buttons_layout.setSpacing(12)
        export_button = QPushButton("📊 Exporter Résultats")
        export_button.setFixedSize(200, 48)
        export_button.clicked.connect(self.export_results)
        delete_button = QPushButton("🗑️ Supprimer Fichier")
        delete_button.setObjectName("deleteButton")
        delete_button.setFixedSize(200, 48)
        delete_button.clicked.connect(self.delete_selected_file)
        close_button = QPushButton("Fermer")
        close_button.setObjectName("cancelButton")
        close_button.setFixedSize(140, 48)
        close_button.clicked.connect(self.close)
        buttons_layout.addWidget(export_button)
        buttons_layout.addWidget(delete_button)
        buttons_layout.addStretch()
        buttons_layout.addWidget(close_button)
        layout.addWidget(buttons_card)
        self.setLayout(layout)
    
    def log_message(self, message):
        """Log message to parent window's log display"""
        if self.parent and hasattr(self.parent, 'log_message'):
            self.parent.log_message(message)
    
    def search_options(self):
        """Search options in database"""
        try:
            search_term = self.search_input.text().strip()
            treatment_code = self.treatment_combo.currentData()
            
            if not search_term:
                self.log_message("Warning: Veuillez entrer un terme de recherche")
                QMessageBox.warning(self, "Warning", "Veuillez entrer un terme de recherche")
                return
            
            self.log_message(f"Recherche de l'option: {search_term}, code de traitement: {treatment_code}")
            results = self.db_manager.search_options(search_term, treatment_code)
            
            self.results_table.setRowCount(0)
            self.results_table.setRowCount(len(results))
            
            for row, result in enumerate(results):
                if len(result) < 8:
                    self.log_message(f"[ERROR] Résultat incomplet à la ligne {row}: {result}")
                    continue
                option_name, procedure_name, is_covered, coverage_details, treatment_code_result, filename, filepath, file_id = result
                self.log_message(f"[DEBUG] Ligne {row}: option_name={option_name}, filename={filename}, filepath={filepath}, file_id={file_id}")
                self.results_table.setItem(row, 0, QTableWidgetItem(option_name or ""))
                self.results_table.setItem(row, 1, QTableWidgetItem(filename or ""))
                self.results_table.setItem(row, 2, QTableWidgetItem(filepath or ""))
                self.results_table.setItem(row, 3, QTableWidgetItem(str(file_id) if file_id else ""))

            
            self.results_table.resizeColumnsToContents()
            self.results_table.horizontalHeader().setMinimumSectionSize(100)
            self.log_message(f"{len(results)} résultats trouvés")
            
            if not results:
                self.log_message("Aucun résultat trouvé pour les critères de recherche donnés")
                QMessageBox.information(self, "Aucun Résultat", "Aucune option trouvée correspondant aux critères de recherche.")
                
        except Exception as e:
            self.log_message(f"Erreur de recherche: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur de recherche: {e}")
    
    def export_results(self):
        """Export search results to CSV with proper column separation"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, 'Exporter Résultats', 'option_search_results.csv', 'CSV Files (*.csv)')
            
            if file_path:
                data = [["Option", "Fichier", "Chemin"]]
                for row in range(self.results_table.rowCount()):
                    row_data = [self.results_table.item(row, col).text() if self.results_table.item(row, col) else "" for col in range(3)]
                    data.append(row_data)
                
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    writer.writerows(data)
                
                self.log_message(f"Résultats exportés vers {file_path}")
                QMessageBox.information(self, "Succès", f"Résultats exportés vers {file_path}")
        except Exception as e:
            self.log_message(f"Erreur d'exportation: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur d'exportation: {e}")


    def delete_selected_file(self):
        """Delete the selected file and its associated options"""
        try:
            selected_rows = self.results_table.selectionModel().selectedRows()
            if not selected_rows:
                self.log_message("Warning: Aucun fichier sélectionné pour la suppression")
                QMessageBox.warning(self, "Warning", "Veuillez sélectionner une ligne à supprimer")
                return

            row = selected_rows[0].row()
            file_id_item = self.results_table.item(row, 3)
            if not file_id_item:
                self.log_message("Erreur: Aucun ID de fichier trouvé pour la ligne sélectionnée")
                QMessageBox.critical(self, "Erreur", "Aucun ID de fichier trouvé pour la ligne sélectionnée")
                return

            file_id = int(file_id_item.text())
            filename = self.results_table.item(row, 1).text()

            reply = QMessageBox.question(
                self,
                "Confirmer la Suppression",
                f"Êtes-vous sûr de vouloir supprimer le fichier '{filename}' et toutes ses options associées ?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                success = self.db_manager.delete_xml_file(file_id)
                if success:
                    self.results_table.removeRow(row)
                    self.log_message(f"Fichier '{filename}' et ses options supprimés avec succès")
                    QMessageBox.information(self, "Succès", f"Fichier '{filename}' supprimé avec succès")
                else:
                    self.log_message(f"Erreur: Fichier '{filename}' non trouvé dans la base de données")
                    QMessageBox.critical(self, "Erreur", f"Fichier '{filename}' non trouvé dans la base de données")
        except Exception as e:
            self.log_message(f"Erreur de suppression: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur de suppression: {e}")


class AdminWindow(QWidget):
    """Fenêtre pour l'administration des utilisateurs, avec interface professionnelle redesignée."""

    def __init__(self, current_user):
        """Initialise la fenêtre d'administration."""
        super().__init__()
        self.current_user = current_user
        self.db_manager = DatabaseManager()
        self.email_manager = EmailManager()
        self.login_window = None  # Référence pour éviter la suppression
        self.email_thread = None  # Référence pour le thread d'email
        self.configure_email_settings()
        self.initUI()
        self.load_users()
        print("[DEBUG] AdminWindow initialized")

    def configure_email_settings(self):
        """Configure les paramètres SMTP pour l'envoi d'emails."""
        self.email_manager.configure_smtp(
            smtp_server="smtp.gmail.com",
            smtp_port=587,
            sender_email="medalilahmar00@gmail.com",  # ⚠️ Remplacez par votre email
            sender_password="tlwu nhmm japb dgnu",  # ⚠️ Remplacez par votre mot de passe d'application
            app_name="PyQt5"
        )
        logo_path = "images/pdflogo.png"
        self.logo_path = logo_path if os.path.exists(logo_path) else None
        print(f"[DEBUG] AdminWindow Logo path: {self.logo_path}")

    def initUI(self):
        """Initialise l'interface utilisateur professionnelle."""
        self.setWindowTitle('🛡️ Administration des Utilisateurs')
        self.setWindowIcon(QIcon("images/iconapp.ico"))
        self.setFixedSize(1700, 970)  # Taille plus grande pour accommoder la grande table
        
        # Assurer que la fenêtre reste visible
        self.setAttribute(Qt.WA_DeleteOnClose, False)
        self.setWindowFlags(Qt.Window | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        
        # Style CSS professionnel et moderne avec table noire dominante
        self.setStyleSheet("""
            QWidget {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #000000,
                    stop: 0.4 #1a1a2e,
                    stop: 1 #0f2f2f
                );
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
            }
            
            /* Header principal - Plus compact */
            #headerFrame {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 rgba(18, 224, 214, 0.15),
                    stop: 0.5 rgba(18, 224, 214, 0.08),
                    stop: 1 rgba(18, 224, 214, 0.15)
                );
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 12px;
                padding: 15px;
                margin-bottom: 15px;
                max-height: 100px;
            }
            
            #mainTitle {
                color: #12e0d6;
                font-size: 24px;
                font-weight: bold;
                text-align: center;
                margin: 0;
                text-shadow: 0 2px 10px rgba(18, 224, 214, 0.3);
            }
            
            #userInfo {
                color: #ffffff;
                font-size: 12px;
                margin-top: 5px;
                text-align: center;
                opacity: 0.9;
            }
            
            /* Table des utilisateurs - ÉNORME ET DOMINANTE */
            QTableWidget {
                background-color: #000000;
                color: #ffffff;
                border: 2px solid #12e0d6;
                border-radius: 12px;
                gridline-color: rgba(18, 224, 214, 0.4);
                selection-background-color: rgba(18, 224, 214, 0.6);
                font-size: 10px;
                font-weight: bold;
                min-height: 400px;
            }
            
            QTableWidget::item {
                padding: 10px 8px;
                border-bottom: 2px solid rgba(18, 224, 214, 0.3);
                color: #ffffff;
                font-size: 14px;
                font-weight: bold;
            }
            
            QTableWidget::item:selected {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 rgba(18, 224, 214, 0.8),
                    stop: 1 rgba(15, 184, 176, 0.8)
                );
                color: #000000;
                font-weight: 900;
                border: 2px solid #12e0d6;
            }
            
            QTableWidget::item:hover {
                background-color: rgba(18, 224, 214, 0.3);
                color: #ffffff;
                font-weight: bold;
            }
            
            QHeaderView::section {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 0.5 #0fb8b0,
                    stop: 1 #0d8f88
                );
                color: #000000;
                font-weight: 900;
                font-size: 18px;
                padding: 25px 15px;
                border: none;
                border-right: 2px solid rgba(0, 0, 0, 0.3);
                text-align: center;
            }
            
            QHeaderView::section:first {
                border-top-left-radius: 15px;
            }
            
            QHeaderView::section:last {
                border-top-right-radius: 15px;
                border-right: none;
            }
            
            /* Statistiques - Plus compact */
            #statsFrame {
                background: rgba(18, 224, 214, 0.1);
                border: 1px solid rgba(18, 224, 214, 0.3);
                border-radius: 10px;
                padding: 10px;
                margin-bottom: 10px;
                max-height: 50px;
            }
            
            #statsLabel {
                color: #12e0d6;
                font-size: 14px;
                font-weight: bold;
            }
            
            /* Boutons d'actions - Plus compact */
            #actionFrame {
                background: rgba(18, 224, 214, 0.05);
                border: 1px solid rgba(18, 224, 214, 0.2);
                border-radius: 15px;
                padding: 15px;
                margin-top: 15px;
                max-height: 150px;
            }
            
            QPushButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 0.5 #0fb8b0,
                    stop: 1 #0d8f88
                );
                color: #000000;
                border: 2px solid rgba(15, 184, 176, 0.8);
                border-radius: 10px;
                padding: 10px 15px;
                font-size: 12px;
                font-weight: bold;
                min-width: 110px;
                min-height: 35px;
            }
            
            QPushButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #0fb8b0,
                    stop: 1 #0d8f88
                );
                border-color: #0d8f88;
                transform: translateY(-2px);
                box-shadow: 0 6px 20px rgba(18, 224, 214, 0.4);
            }
            
            QPushButton:pressed {
                background: #0d8f88;
                transform: translateY(0px);
                box-shadow: 0 2px 10px rgba(18, 224, 214, 0.3);
            }
            
            QPushButton:disabled {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #4a5568,
                    stop: 1 #2d3748
                );
                color: #a0aec0;
                border-color: #4a5568;
            }
            
            /* Boutons spécialisés */
            #activateBtn {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #10b981,
                    stop: 1 #059669
                );
                border-color: #059669;
            }
            
            #activateBtn:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #059669,
                    stop: 1 #047857
                );
                box-shadow: 0 6px 20px rgba(16, 185, 129, 0.4);
            }
            
            #deactivateBtn {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #ef4444,
                    stop: 1 #dc2626
                );
                border-color: #dc2626;
            }
            
            #deactivateBtn:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #dc2626,
                    stop: 1 #b91c1c
                );
                box-shadow: 0 6px 20px rgba(239, 68, 68, 0.4);
            }
            
            #unlockBtn {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #f59e0b,
                    stop: 1 #d97706
                );
                border-color: #d97706;
            }
            
            #unlockBtn:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #d97706,
                    stop: 1 #b45309
                );
                box-shadow: 0 6px 20px rgba(245, 158, 11, 0.4);
            }
            
            #deleteBtn {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #9b2c2c,
                    stop: 1 #7f1d1d
                );
                border-color: #7f1d1d;
            }
            
            #deleteBtn:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #7f1d1d,
                    stop: 1 #631718
                );
                box-shadow: 0 6px 20px rgba(155, 44, 44, 0.4);
            }
            
            #backBtn {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #6b7280,
                    stop: 1 #4b5563
                );
                border-color: #4b5563;
            }
            
            #backBtn:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #4b5563,
                    stop: 1 #374151
                );
                box-shadow: 0 6px 20px rgba(107, 114, 128, 0.4);
            }
            
            /* Labels et statut - Plus compact */
            QLabel {
                color: #ffffff;
                font-size: 12px;
                background: transparent;
            }
            
            #statusLabel {
                color: #12e0d6;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                background: rgba(18, 224, 214, 0.1);
                border: 1px solid rgba(18, 224, 214, 0.3);
                border-radius: 6px;
                margin-top: 8px;
                max-height: 40px;
            }
            
            /* Scrollbar personnalisée pour la grande table */
            QScrollBar:vertical {
                background: rgba(0, 0, 0, 0.8);
                width: 15px;
                border-radius: 7px;
                border: 1px solid #12e0d6;
            }
            
            QScrollBar::handle:vertical {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                border-radius: 7px;
                min-height: 30px;
                border: 1px solid #0d8f88;
            }
            
            QScrollBar::handle:vertical:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #0fb8b0,
                    stop: 1 #0d8f88
                );
            }
        """)
        
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Header avec informations - COMPACT
        header_frame = QFrame()
        header_frame.setObjectName("headerFrame")
        header_layout = QVBoxLayout(header_frame)
        
        title = QLabel('🛡️ Panneau d\'Administration')
        title.setObjectName("mainTitle")
        title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title)
        
        user_info = QLabel(f'👨‍💼 {self.current_user} • 🕐 {datetime.now().strftime("%d/%m/%Y - %H:%M")}')
        user_info.setObjectName("userInfo")
        user_info.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(user_info)
        
        main_layout.addWidget(header_frame)
        
        # Frame des statistiques - COMPACT
        stats_frame = QFrame()
        stats_frame.setObjectName("statsFrame")
        stats_layout = QHBoxLayout(stats_frame)
        
        self.stats_label = QLabel('📊 Chargement des statistiques...')
        self.stats_label.setObjectName("statsLabel")
        stats_layout.addWidget(self.stats_label)
        
        main_layout.addWidget(stats_frame)
        
        # Table des utilisateurs avec titre - ÉNORME ET DOMINANTE
        table_title = QLabel('👥 GESTION DES UTILISATEURS')
        table_title.setStyleSheet("font-size: 20px; font-weight: 900; color: #12e0d6; margin-bottom: 10px; text-align: center; text-shadow: 0 2px 10px rgba(18, 224, 214, 0.5);")
        table_title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(table_title)
        
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(8)
        self.users_table.setHorizontalHeaderLabels([
            '🆔 ID', '👤 UTILISATEUR', '📧 EMAIL', '🎭 RÔLE', 
            '✅ ACTIF', '🔒 VERROUILLÉ', '⚠️ TENTATIVES', '📅 CRÉÉ LE'
        ])
        
        # Configuration de la table - ÉNORME
        self.users_table.horizontalHeader().setStretchLastSection(True)
        self.users_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.users_table.setAlternatingRowColors(False)  # Désactivé pour un fond noir uniforme
        self.users_table.setSortingEnabled(True)
        self.users_table.setMinimumHeight(500)  # Hauteur minimum énorme
        
        # Ajustement des colonnes - PLUS LARGES
        self.users_table.setColumnWidth(0, 80)    # ID
        self.users_table.setColumnWidth(1, 200)   # Utilisateur
        self.users_table.setColumnWidth(2, 280)   # Email
        self.users_table.setColumnWidth(3, 150)   # Rôle
        self.users_table.setColumnWidth(4, 120)   # Actif
        self.users_table.setColumnWidth(5, 140)   # Verrouillé
        self.users_table.setColumnWidth(6, 140)   # Tentatives
        
        # Hauteur des lignes - PLUS HAUTES
        self.users_table.verticalHeader().setDefaultSectionSize(60)
        self.users_table.verticalHeader().setVisible(False)
        
        main_layout.addWidget(self.users_table)
        
        # Frame des actions - COMPACT
        action_frame = QFrame()
        action_frame.setObjectName("actionFrame")
        action_layout = QVBoxLayout(action_frame)
        
        # Label des actions
        action_title = QLabel('⚡ Actions Administrateur')
        action_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #12e0d6; margin-bottom: 10px;")
        action_layout.addWidget(action_title)
        
        # Première ligne de boutons
        button_layout1 = QHBoxLayout()
        button_layout1.setSpacing(10)
        
        self.activate_button = QPushButton('✅ Activer')
        self.activate_button.setObjectName("activateBtn")
        self.activate_button.clicked.connect(self.activate_user)
        self.activate_button.setToolTip("Active l'utilisateur sélectionné et envoie un email de confirmation")
        button_layout1.addWidget(self.activate_button)
        
        self.deactivate_button = QPushButton('❌ Désactiver')
        self.deactivate_button.setObjectName("deactivateBtn")
        self.deactivate_button.clicked.connect(self.deactivate_user)
        self.deactivate_button.setToolTip("Désactive l'utilisateur sélectionné")
        button_layout1.addWidget(self.deactivate_button)
        
        self.unlock_button = QPushButton('🔓 Déverrouiller')
        self.unlock_button.setObjectName("unlockBtn")
        self.unlock_button.clicked.connect(self.unlock_user)
        self.unlock_button.setToolTip("Déverrouille l'utilisateur bloqué")
        button_layout1.addWidget(self.unlock_button)
        
        self.reset_password_button = QPushButton('🔑 Reset MDP')
        self.reset_password_button.clicked.connect(self.reset_password)
        self.reset_password_button.setToolTip("Réinitialise le mot de passe de l'utilisateur")
        button_layout1.addWidget(self.reset_password_button)
        
        action_layout.addLayout(button_layout1)
        
        # Deuxième ligne de boutons
        button_layout2 = QHBoxLayout()
        button_layout2.setSpacing(10)
        
        self.change_role_button = QPushButton('🎭 Changer Rôle')
        self.change_role_button.clicked.connect(self.change_role)
        self.change_role_button.setToolTip("Modifie le rôle de l'utilisateur (user/admin/readonly)")
        button_layout2.addWidget(self.change_role_button)
        
        self.delete_button = QPushButton('🗑️ Supprimer')
        self.delete_button.setObjectName("deleteBtn")
        self.delete_button.clicked.connect(self.delete_user)
        self.delete_button.setToolTip("Supprime l'utilisateur sélectionné de la base de données")
        button_layout2.addWidget(self.delete_button)
        
        self.refresh_button = QPushButton('🔄 Actualiser')
        self.refresh_button.clicked.connect(self.load_users)
        self.refresh_button.setToolTip("Actualise la liste des utilisateurs")
        button_layout2.addWidget(self.refresh_button)
        
        # Spacer pour centrer le bouton retour
        button_layout2.addStretch()
        
        self.back_button = QPushButton('🏠 Retour')
        self.back_button.setObjectName("backBtn")
        self.back_button.clicked.connect(self.back_to_login)
        self.back_button.setToolTip("Retour au menu principal")
        button_layout2.addWidget(self.back_button)
        
        action_layout.addLayout(button_layout2)
        
        main_layout.addWidget(action_frame)
        
        # Label de statut - COMPACT
        self.status_label = QLabel('')
        self.status_label.setObjectName("statusLabel")
        self.status_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(self.status_label)
        
        self.setLayout(main_layout)
        
        # S'assurer que la fenêtre est visible
        self.show()
        self.raise_()
        self.activateWindow()
    
    def closeEvent(self, event):
        """Gère la fermeture de la fenêtre."""
        print("[DEBUG] AdminWindow: Close event triggered")
        event.accept()
    
    def update_stats(self):
        """Met à jour les statistiques des utilisateurs."""
        try:
            users = self.db_manager.get_all_users()
            total_users = len(users)
            active_users = sum(1 for user in users if user[4])  # Colonne 4 = actif
            locked_users = sum(1 for user in users if user[5])  # Colonne 5 = verrouillé
            admin_users = sum(1 for user in users if user[3] == 'admin')  # Colonne 3 = rôle
            
            stats_text = f'👥 Total: {total_users} • ✅ Actifs: {active_users} • 🔒 Verrouillés: {locked_users} • 🛡️ Admins: {admin_users}'
            self.stats_label.setText(stats_text)
        except Exception as e:
            self.stats_label.setText('⚠️ Erreur lors du chargement des statistiques')
            print(f"[ERROR] Erreur stats: {e}")
    
    def back_to_login(self):
        """Méthode pour revenir à la fenêtre de connexion avec animation."""
        print("[DEBUG] AdminWindow: Navigating back to LoginWindow")
        self.status_label.setText('🔄 Retour au menu principal...')
        QApplication.processEvents()  # Force l'affichage du message
        
        # Import local pour éviter les imports circulaires
        try:
            self.login_window = LoginWindow()
            self.login_window.show()
            self.hide()  # Cacher au lieu de fermer
        except ImportError:
            print("[ERROR] Impossible d'importer LoginWindow")
            self.close()
    
    def load_users(self):
        """Charge les utilisateurs dans la table avec mise en forme améliorée."""
        try:
            self.status_label.setText('🔄 Chargement des utilisateurs...')
            QApplication.processEvents()
            
            users = self.db_manager.get_all_users()
            self.users_table.setRowCount(len(users))
            
            for row, user in enumerate(users):
                for col, value in enumerate(user):
                    if col == 4:  # Actif
                        value = "🟢 OUI" if value else "🔴 NON"
                    elif col == 5:  # Verrouillé
                        value = "🔒 OUI" if value else "🔓 NON"
                    elif col == 3:  # Rôle
                        role_icons = {'admin': '🛡️', 'user': '👤', 'readonly': '👁️'}
                        icon = role_icons.get(str(value), '❓')
                        value = f"{icon} {value.upper()}"
                    elif col == 6:  # Tentatives
                        if value and int(value) > 0:
                            value = f"⚠️ {value}"
                        else:
                            value = "✅ 0"
                    elif col == 1:  # Utilisateur
                        value = f"👤 {value}"
                    elif col == 2:  # Email
                        value = f"📧 {value}"
                    
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    item.setTextAlignment(Qt.AlignCenter)
                    
                    # Coloration conditionnelle sur fond noir
                    if col == 4:  # Colonne Actif
                        if "🟢" in str(value):
                            item.setBackground(QColor(0, 100, 0, 100))  # Vert foncé transparent
                            item.setForeground(QColor(144, 238, 144))   # Texte vert clair
                        else:
                            item.setBackground(QColor(100, 0, 0, 100))  # Rouge foncé transparent
                            item.setForeground(QColor(255, 182, 193))   # Texte rouge clair
                    elif col == 5:  # Colonne Verrouillé
                        if "🔒" in str(value):
                            item.setBackground(QColor(100, 50, 0, 100))  # Orange foncé transparent
                            item.setForeground(QColor(255, 215, 0))      # Texte jaune/orange clair
                        else:
                            item.setForeground(QColor(144, 238, 144))    # Texte vert clair
                    elif col == 3:  # Colonne Rôle
                        if "🛡️" in str(value):
                            item.setBackground(QColor(0, 50, 50, 100))   # Cyan foncé transparent
                            item.setForeground(QColor(18, 224, 214))     # Texte cyan brillant
                        else:
                            item.setForeground(QColor(255, 255, 255))    # Texte blanc
                    elif col == 6:  # Colonne Tentatives
                        if "⚠️" in str(value):
                            item.setForeground(QColor(255, 215, 0))      # Texte jaune
                        else:
                            item.setForeground(QColor(144, 238, 144))    # Texte vert clair
                    else:
                        item.setForeground(QColor(255, 255, 255))        # Texte blanc par défaut
                    
                    # Police en gras pour tous les éléments
                    font = item.font()
                    font.setBold(True)
                    font.setPointSize(14)
                    item.setFont(font)
                    
                    self.users_table.setItem(row, col, item)
            
            self.update_stats()
            self.status_label.setText(f'✅ {len(users)} utilisateurs chargés avec succès')
            print(f"[DEBUG] AdminWindow: Loaded {len(users)} users")
            
        except Exception as e:
            self.status_label.setText('❌ Erreur lors du chargement')
            print(f"[ERROR] Erreur chargement utilisateurs: {e}")
    
    def get_selected_user_id(self):
        """Récupère l'ID de l'utilisateur sélectionné."""
        current_row = self.users_table.currentRow()
        if current_row >= 0:
            return int(self.users_table.item(current_row, 0).text())
        QMessageBox.warning(self, '⚠️ Sélection requise', 
                           'Veuillez sélectionner un utilisateur dans la liste.')
        return None
    
    def get_selected_username(self):
        """Récupère le nom d'utilisateur sélectionné."""
        current_row = self.users_table.currentRow()
        if current_row >= 0:
            username_text = self.users_table.item(current_row, 1).text()
            # Enlever l'icône 👤 du début
            return username_text.replace("👤 ", "") if "👤 " in username_text else username_text
        return None
    
    def get_selected_email(self):
        """Récupère l'email de l'utilisateur sélectionné."""
        current_row = self.users_table.currentRow()
        if current_row >= 0:
            email_text = self.users_table.item(current_row, 2).text()
            # Enlever l'icône 📧 du début
            return email_text.replace("📧 ", "") if "📧 " in email_text else email_text
        return None
    
    def activate_user(self):
        """Active l'utilisateur sélectionné et envoie un email d'activation."""
        user_id = self.get_selected_user_id()
        if user_id:
            username = self.get_selected_username()
            email = self.get_selected_email()
            print(f"[DEBUG] AdminWindow: Activating user id={user_id}, username={username}, email={email}")
            
            if username and email:
                # Confirmation avant activation
                reply = QMessageBox.question(
                    self, 
                    '✅ Confirmer l\'activation', 
                    f'Êtes-vous sûr de vouloir activer l\'utilisateur:\n\n👤 {username}\n📧 {email}\n\nUn email de confirmation sera envoyé.',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    self.db_manager.update_user_status(user_id, active=1)
                    self.status_label.setText('📧 Envoi de l\'email de validation...')
                    self.send_activation_email(email, username)
            else:
                self.status_label.setText('')
                QMessageBox.warning(self, '❌ Erreur', 
                                  'Nom d\'utilisateur ou email non trouvé.')
    
    def send_activation_email(self, recipient_email, username):
        """Envoie un email d'activation dans un thread séparé."""
        print(f"[DEBUG] AdminWindow: Sending activation email to {recipient_email}")
        logo_path = self.logo_path if self.logo_path and os.path.exists(self.logo_path) else None
        
        # Import local pour éviter les imports circulaires
        try:
            self.email_thread = EmailSendThread(
                self.email_manager,
                recipient_email,
                username,
                cover_image_path="images/neoxam.png",
                email_type="acceptance"
            )
            self.email_thread.email_sent.connect(self.on_activation_email_sent)
            self.email_thread.start()
        except ImportError as e:
            print(f"[ERROR] Impossible d'importer EmailSendThread: {e}")
            self.status_label.setText('⚠️ Erreur lors de l\'envoi d\'email')
    
    def on_activation_email_sent(self, success, message):
        """Gère le résultat de l'envoi de l'email d'activation."""
        print(f"[DEBUG] AdminWindow: Activation email result - success={success}, message={message}")
        self.load_users()
        
        if success:
            self.status_label.setText('✅ Utilisateur activé et email envoyé!')
            QMessageBox.information(
                self, 
                '✅ Succès', 
                '🎉 Utilisateur activé avec succès!\n\n📧 Email de validation envoyé.'
            )
        else:
            self.status_label.setText('⚠️ Activé, erreur email')
            QMessageBox.information(
                self, 
                '⚠️ Partiellement réussi', 
                f'✅ Utilisateur activé avec succès!\n\n⚠️ Email non envoyé:\n{message}'
            )
    
    def deactivate_user(self):
        """Désactive l'utilisateur sélectionné avec confirmation."""
        user_id = self.get_selected_user_id()
        if user_id:
            username = self.get_selected_username()
            
            # Confirmation avant désactivation
            reply = QMessageBox.question(
                self,
                '❌ Confirmer la désactivation',
                f'Êtes-vous sûr de vouloir désactiver l\'utilisateur:\n\n👤 {username}\n\nL\'utilisateur ne pourra plus se connecter.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.status_label.setText('🔄 Désactivation en cours...')
                self.db_manager.update_user_status(user_id, active=0)
                self.load_users()
                self.status_label.setText('✅ Utilisateur désactivé')
                QMessageBox.information(self, '✅ Succès', 
                                      f'👤 {username} a été désactivé avec succès.')
    
    def unlock_user(self):
        """Déverrouille l'utilisateur sélectionné."""
        user_id = self.get_selected_user_id()
        if user_id:
            username = self.get_selected_username()
            
            reply = QMessageBox.question(
                self,
                '🔓 Confirmer le déverrouillage',
                f'Déverrouiller l\'utilisateur:\n\n👤 {username}\n\nLes tentatives de connexion seront remises à zéro.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.status_label.setText('🔄 Déverrouillage en cours...')
                self.db_manager.update_user_status(user_id, locked=0)
                self.load_users()
                self.status_label.setText('✅ Utilisateur déverrouillé')
                QMessageBox.information(self, '✅ Succès', 
                                      f'🔓 {username} a été déverrouillé avec succès.')
    
    def delete_user(self):
        """Supprime l'utilisateur sélectionné avec confirmation."""
        user_id = self.get_selected_user_id()
        if user_id:
            username = self.get_selected_username()
            
            # Confirmation avant suppression
            reply = QMessageBox.question(
                self,
                '🗑️ Confirmer la suppression',
                f'Êtes-vous sûr de vouloir supprimer l\'utilisateur:\n\n👤 {username}\n\nCette action est irréversible.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.status_label.setText('🔄 Suppression en cours...')
                success, message = self.db_manager.delete_user(user_id)
                if success:
                    self.load_users()
                    self.status_label.setText('✅ Utilisateur supprimé')
                    QMessageBox.information(
                        self, 
                        '✅ Succès', 
                        f'🗑️ {username} a été supprimé avec succès.'
                    )
                else:
                    self.status_label.setText('❌ Erreur lors de la suppression')
                    QMessageBox.warning(
                        self, 
                        '⚠️ Erreur', 
                        f'Impossible de supprimer l\'utilisateur:\n{message}'
                    )
    
    def reset_password(self):
        """Réinitialise le mot de passe de l'utilisateur sélectionné."""
        user_id = self.get_selected_user_id()
        if user_id:
            username = self.get_selected_username()
            
            # Boîte de dialogue personnalisée pour le nouveau mot de passe
            dialog = QDialog(self)
            dialog.setWindowTitle('🔑 Réinitialiser le mot de passe')
            self.setWindowIcon(QIcon("images/iconapp.ico"))
            dialog.setFixedSize(500, 500)
            dialog.setStyleSheet("""
                QDialog {
                    background: qlineargradient(
                        x1: 0, y1: 0, x2: 1, y2: 1,
                        stop: 0 #1a1a2e,
                        stop: 1 #0f2f2f
                    );
                    border-radius: 15px;
                }
                QLabel {
                    color: #ffffff;
                    font-size: 14px;
                    font-weight: bold;
                }
                QLineEdit {
                    background-color: #1C2526;
                    color: #FFFFFF;
                    border: 2px solid rgba(18, 224, 214, 0.3);
                    border-radius: 8px;
                    padding: 10px;
                    font-size: 14px;
                }
                QLineEdit:focus {
                    border: 2px solid #0fb8b0;
                }
                QPushButton {
                    background-color: #12e0d6;
                    color: #000000;
                    border: 2px solid #0fb8b0;
                    border-radius: 8px;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: bold;
                    min-width: 100px;
                }
                QPushButton:hover {
                    background-color: #0fb8b0;
                }
            """)
            
            layout = QVBoxLayout(dialog)
            layout.setSpacing(15)
            layout.setContentsMargins(20, 20, 20, 20)
            
            title_label = QLabel(f'🔑 Nouveau mot de passe pour:\n👤 {username}')
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet("font-size: 16px; color: #12e0d6; margin-bottom: 10px;")
            layout.addWidget(title_label)
            
            password_label = QLabel('Nouveau mot de passe:')
            layout.addWidget(password_label)
            
            password_input = QLineEdit()
            password_input.setEchoMode(QLineEdit.Password)
            password_input.setPlaceholderText('Minimum 6 caractères...')
            layout.addWidget(password_input)
            
            confirm_label = QLabel('Confirmer le mot de passe:')
            layout.addWidget(confirm_label)
            
            confirm_input = QLineEdit()
            confirm_input.setEchoMode(QLineEdit.Password)
            confirm_input.setPlaceholderText('Confirmer le mot de passe...')
            layout.addWidget(confirm_input)
            
            button_layout = QHBoxLayout()
            ok_button = QPushButton('✅ Confirmer')
            cancel_button = QPushButton('❌ Annuler')
            
            button_layout.addWidget(ok_button)
            button_layout.addWidget(cancel_button)
            layout.addLayout(button_layout)
            
            def validate_and_accept():
                password = password_input.text()
                confirm = confirm_input.text()
                
                if len(password) < 6:
                    QMessageBox.warning(dialog, '⚠️ Erreur', 
                                      'Le mot de passe doit contenir au moins 6 caractères.')
                    return
                
                if password != confirm:
                    QMessageBox.warning(dialog, '⚠️ Erreur', 
                                      'Les mots de passe ne correspondent pas.')
                    return
                
                dialog.new_password = password
                dialog.accept()
            
            ok_button.clicked.connect(validate_and_accept)
            cancel_button.clicked.connect(dialog.reject)
            password_input.returnPressed.connect(validate_and_accept)
            confirm_input.returnPressed.connect(validate_and_accept)
            
            if dialog.exec_() == QDialog.Accepted:
                self.status_label.setText('🔄 Réinitialisation du mot de passe...')
                success, message = self.db_manager.reset_password(user_id, dialog.new_password)
                if success:
                    self.status_label.setText('✅ Mot de passe réinitialisé')
                    QMessageBox.information(
                        self, 
                        '✅ Succès', 
                        f'🔑 Mot de passe réinitialisé pour {username}'
                    )
                else:
                    self.status_label.setText('❌ Erreur lors de la réinitialisation')
                    QMessageBox.warning(
                        self, 
                        '⚠️ Erreur', 
                        f'Impossible de réinitialiser le mot de passe:\n{message}'
                    )
    
    def change_role(self):
        """Change le rôle de l'utilisateur sélectionné avec interface améliorée."""
        user_id = self.get_selected_user_id()
        if user_id:
            username = self.get_selected_username()
            current_role = self.users_table.item(self.users_table.currentRow(), 3).text()
            
            # Nettoyer le rôle actuel (enlever l'icône et convertir en minuscules)
            current_role_clean = current_role.split(' ')[-1].lower() if ' ' in current_role else current_role.lower()
            
            # Boîte de dialogue personnalisée pour changer le rôle
            dialog = QDialog(self)
            dialog.setWindowTitle('🎭 Changer le rôle')
            self.setWindowIcon(QIcon("images/iconapp.ico"))
            dialog.setFixedSize(500, 500)
            dialog.setStyleSheet("""
                QDialog {
                    background: qlineargradient(
                        x1: 0, y1: 0, x2: 1, y2: 1,
                        stop: 0 #1a1a2e,
                        stop: 1 #0f2f2f
                    );
                    border-radius: 15px;
                }
                QLabel {
                    color: #ffffff;
                    font-size: 14px;
                    font-weight: bold;
                }
                QRadioButton {
                    color: #ffffff;
                    font-size: 14px;
                    padding: 8px;
                    margin: 5px;
                }
                QRadioButton::indicator {
                    width: 18px;
                    height: 18px;
                }
                QRadioButton::indicator:unchecked {
                    background-color: #1C2526;
                    border: 2px solid rgba(18, 224, 214, 0.3);
                    border-radius: 9px;
                }
                QRadioButton::indicator:checked {
                    background-color: #12e0d6;
                    border: 2px solid #0fb8b0;
                    border-radius: 9px;
                }
                QPushButton {
                    background-color: #12e0d6;
                    color: #000000;
                    border: 2px solid #0fb8b0;
                    border-radius: 8px;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: bold;
                    min-width: 100px;
                }
                QPushButton:hover {
                    background-color: #0fb8b0;
                }
            """)
            
            layout = QVBoxLayout(dialog)
            layout.setSpacing(15)
            layout.setContentsMargins(20, 20, 20, 20)
            
            title_label = QLabel(f'🎭 Changer le rôle de:\n👤 {username}')
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet("font-size: 16px; color: #12e0d6; margin-bottom: 15px;")
            layout.addWidget(title_label)
            
            current_label = QLabel(f'Rôle actuel: {current_role}')
            current_label.setAlignment(Qt.AlignCenter)
            current_label.setStyleSheet("color: #f59e0b; margin-bottom: 15px;")
            layout.addWidget(current_label)
            
            # Options de rôles avec descriptions
            roles_info = {
                'user': ('👤 Utilisateur', 'Accès standard aux fonctionnalités'),
                'admin': ('🛡️ Administrateur', 'Accès complet à toutes les fonctions'),
                'readonly': ('👁️ Lecture seule', 'Accès en lecture uniquement')
            }
            
            role_group = QButtonGroup(dialog)
            role_buttons = {}
            
            for role, (display, description) in roles_info.items():
                radio = QRadioButton(f'{display}\n{description}')
                radio.setStyleSheet("padding: 10px; margin: 5px;")
                if role == current_role_clean:
                    radio.setChecked(True)
                role_group.addButton(radio)
                role_buttons[radio] = role
                layout.addWidget(radio)
            
            button_layout = QHBoxLayout()
            ok_button = QPushButton('✅ Appliquer')
            cancel_button = QPushButton('❌ Annuler')
            
            button_layout.addWidget(ok_button)
            button_layout.addWidget(cancel_button)
            layout.addLayout(button_layout)
            
            def apply_role_change():
                selected_button = role_group.checkedButton()
                if selected_button:
                    new_role = role_buttons[selected_button]
                    if new_role != current_role_clean:
                        reply = QMessageBox.question(
                            dialog,
                            '🎭 Confirmer le changement',
                            f'Changer le rôle de {username}:\n\n'
                            f'De: {current_role}\n'
                            f'Vers: {roles_info[new_role][0]}\n\n'
                            f'Êtes-vous sûr ?',
                            QMessageBox.Yes | QMessageBox.No,
                            QMessageBox.No
                        )
                        if reply == QMessageBox.Yes:
                            dialog.new_role = new_role
                            dialog.accept()
                    else:
                        QMessageBox.information(dialog, 'ℹ️ Information', 
                                              'Le rôle sélectionné est déjà le rôle actuel.')
            
            ok_button.clicked.connect(apply_role_change)
            cancel_button.clicked.connect(dialog.reject)
            
            if dialog.exec_() == QDialog.Accepted:
                self.status_label.setText('🔄 Changement de rôle en cours...')
                self.db_manager.update_user_status(user_id, role=dialog.new_role)
                self.load_users()
                self.status_label.setText('✅ Rôle modifié avec succès')
                QMessageBox.information(
                    self,
                    '✅ Succès',
                    f'🎭 Rôle de {username} changé vers:\n{roles_info[dialog.new_role][0]}'
                )    



class XMLParser:
    """Enhanced XML parser with improved validation and coverage checking"""

    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.required_fields = {
            'set_variable': ['name', 'value', 'id', 'code', 'date_debut', 'slo_objet'],
            'field_input': ['name', 'value', 'input_', 'required_', 'DATE_DEBUT', 'SLO_OBJET'],
            'option': ['name', 'prog_name']
        }
        self.coverage_rules = {
            'mandatory_attributes': ['name', 'prog_name'],
            'minimum_variables': 1
        }

    def parse_xml_file(self, filepath, log_func=None):
        """Parse XML file with enhanced validation and logging"""
        try:
            tree = ET.parse(filepath)
            root = tree.getroot()

            file_hash = self._calculate_file_hash(filepath)
            file_id = self.db_manager.add_xml_file(
                os.path.basename(filepath),
                filepath,
                file_hash
            )

            parsed_data = {
                'file_id': file_id,
                'options': [],
                'variables': [],
                'treatment_codes': set(),
                'filepath': filepath  # Added filepath for consistency
            }

            processed_options = set()  # Track processed options to avoid duplicates
            option_count = 0
            variable_count = 0

            for element in root.iter():
                if element.tag == 'option':
                    option_name = element.get('name')
                    prog_name = element.get('prog_name') or ''  # Handle None case

                    if option_name and option_name not in {'MAJPROG', 'MAJMENU', 'FORPCT', 'MAJOPT'}:
                        option_key = (file_id, option_name, prog_name)
                        if option_key not in processed_options:
                            processed_options.add(option_key)
                            coverage_details = self._check_coverage(element)
                            current_option = {
                                'name': option_name,
                                'procedure': prog_name,
                                'is_covered': coverage_details['is_covered'],
                                'coverage_details': coverage_details['details'],
                                'treatment_code': prog_name,
                                'variables': []
                            }
                            option_id = self.db_manager.add_option(
                                file_id,
                                option_name,
                                prog_name,
                                coverage_details['is_covered'],
                                prog_name,
                                coverage_details['details']
                            )
                            parsed_data['options'].append(current_option)
                            option_count += 1

                            if prog_name:
                                parsed_data['treatment_codes'].add(prog_name)
                                self.db_manager.add_treatment_code(prog_name)

                            # Parse variables for this option
                            for child in element:
                                if child.tag in ['set_variable', 'field_input']:
                                    var_name = child.get('name')
                                    var_value = child.get('value', '')
                                    # Ensure valid variable name and value
                                    if var_name and var_value and var_value not in {'@SKIP@', '@DOWN@', '@QUIT@'}:
                                        if var_name not in {'code_option', 'nom_procedure', 'nature_option_saisie'}:
                                            variable = {
                                                'name': var_name,
                                                'value': var_value,
                                                'type': child.tag,
                                                'is_required': self._is_required_field(var_name, child.tag)
                                            }
                                            current_option['variables'].append(variable)
                                            parsed_data['variables'].append(variable)
                                            self.db_manager.add_variable(
                                                option_id,
                                                var_name,
                                                var_value,
                                                child.tag,
                                                variable['is_required']
                                            )
                                            variable_count += 1
                                            if log_func:
                                                log_func(f"Parsed {child.tag}: {var_name} = {var_value}, is_required: {variable['is_required']}")

            if log_func:
                log_func(f"Parsing complete: {option_count} options, {variable_count} variables processed")

            return parsed_data

        except ET.ParseError as e:
            if log_func:
                log_func(f"XML parsing error: {e}")
            print(f"XML parsing error: {e}")
            return None
        except sqlite3.Error as e:
            if log_func:
                log_func(f"Database error during XML parsing: {e}")
            print(f"Database error during XML parsing: {e}")
            return None
        except Exception as e:
            if log_func:
                log_func(f"Unexpected error during XML parsing: {e}")
            print(f"Unexpected error during XML parsing: {e}")
            return None

    def _calculate_file_hash(self, filepath):
        """Calculate MD5 hash of file"""
        try:
            hash_md5 = hashlib.md5()
            with open(filepath, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            print(f"Error calculating file hash: {e}")
            raise

    def _check_coverage(self, element):
        """Enhanced coverage checking"""
        coverage_details = {'is_covered': True, 'details': []}

        # Check mandatory attributes
        for attr in self.coverage_rules['mandatory_attributes']:
            if not element.get(attr):
                coverage_details['is_covered'] = False
                coverage_details['details'].append(f"Missing mandatory attribute: {attr}")

        # Check minimum variable count
        variable_count = sum(1 for child in element if child.tag in ['set_variable', 'field_input'])
        if variable_count < self.coverage_rules['minimum_variables']:
            coverage_details['is_covered'] = False
            coverage_details['details'].append(f"Insufficient variables: {variable_count}/{self.coverage_rules['minimum_variables']}")

        return {
            'is_covered': coverage_details['is_covered'],
            'details': '; '.join(coverage_details['details']) if coverage_details['details'] else 'Fully covered'
        }

    def _is_required_field(self, field_name, field_type):
        """Determine if field is required"""
        required_patterns = {
            'set_variable': [r'^id$', r'^name$', r'^code$', r'^date_debut$', r'^slo_objet$'],
            'field_input': [r'^input_$', r'^required_$', r'^DATE_DEBUT$', r'^SLO_OBJET$']
        }
        return any(re.match(pattern, field_name, re.IGNORECASE) for pattern in required_patterns[field_type])






class ReportGenerator:
    """Class for generating enhanced PDF, CSV, and Excel reports"""
    def __init__(self, parsed_data, nomenclature, log_func=None):
        self.parsed_data = parsed_data
        self.nomenclature = nomenclature
        self.log_func = log_func
        if self.log_func:
            self.log_func(f"Initialized ReportGenerator with nomenclature: {nomenclature}")

    def generate_pdf(self, type_etat=1):
        """Generate enhanced PDF report with consistent Neoxam design"""
        try:
            if not self.parsed_data.get('options'):
                if self.log_func:
                    self.log_func("No options available for PDF report")
                return None

            os.makedirs("pdfs", exist_ok=True)
            pdf_path = os.path.join("pdfs", f"{self.nomenclature}.pdf")
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=A4,
                rightMargin=50,
                leftMargin=50,
                topMargin=100,
                bottomMargin=80,
                canvasmaker=NumberedCanvas
            )

            styles = getSampleStyleSheet()
            neoxam_green = colors.HexColor('#00C4B4')
            neoxam_dark = colors.HexColor('#0F172A')
            neoxam_gray = colors.HexColor('#4B5563')
            neoxam_light_gray = colors.HexColor('#F8F9FA')

            title_style = ParagraphStyle(
                'NeoxamTitle',
                parent=styles['Heading1'],
                fontSize=28,
                spaceAfter=30,
                textColor=neoxam_green,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            subtitle_style = ParagraphStyle(
                'NeoxamSubtitle',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=15,
                textColor=neoxam_dark,
                alignment=TA_LEFT,
                fontName='Helvetica-Bold'
            )

            normal_style = ParagraphStyle(
                'NeoxamNormal',
                parent=styles['Normal'],
                fontSize=11,
                textColor=neoxam_gray,
                fontName='Helvetica',
                spaceAfter=10,
                leading=14
            )

            table_cell_style = ParagraphStyle(
                'TableCell',
                parent=styles['Normal'],
                fontSize=10,
                textColor=neoxam_dark,
                fontName='Helvetica',
                leading=12,
                wordWrap='CJK'
            )

            elements = []

            logo = create_neoxam_logo_placeholder()
            if logo:
                logo.hAlign = 'LEFT'
                elements.append(logo)
            else:
                logo_style = ParagraphStyle(
                    'LogoStyle',
                    parent=styles['Normal'],
                    fontSize=24,
                    textColor=neoxam_green,
                    fontName='Helvetica-Bold',
                    alignment=TA_LEFT
                )
                logo_text = Paragraph("🏢 NEOXAM", logo_style)
                elements.append(logo_text)

            elements.append(Spacer(1, 20))

            title = Paragraph("RAPPORT D'ANALYSE XML", title_style)
            elements.append(title)

            summary_title = Paragraph("📊 Résumé Exécutif", subtitle_style)
            elements.append(summary_title)

            total_options = len(self.parsed_data.get('options', []))
            total_variables = len(self.parsed_data.get('variables', []))
            unique_options = len(set(option.get('name', '') for option in self.parsed_data.get('options', []) if option.get('name') and isinstance(option.get('name'), str)))

            summary_data = [
                ["📄 Nom du fichier", os.path.basename(self.parsed_data.get('filepath', 'Unknown'))],
                ["🔢 Nombre d'options", str(total_options)],
                ["📋 Nombre de variables", str(total_variables)],
                ["🔍 Nombre d'options uniques", str(unique_options)],
                ["📅 Date de génération", datetime.now().strftime('%d/%m/%Y')],
                ["⏰ Heure de génération", datetime.now().strftime('%H:%M:%S')]
            ]

            summary_table = Table(summary_data, colWidths=[2.5*inch, 3.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), neoxam_green),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (0, -1), 11),
                ('BACKGROUND', (1, 0), (1, -1), neoxam_light_gray),
                ('TEXTCOLOR', (1, 0), (1, -1), neoxam_dark),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (1, 0), (1, -1), 11),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LINEBELOW', (0, 0), (-1, -1), 1, colors.HexColor('#E9ECEF')),
                ('LINEAFTER', (0, 0), (0, -1), 2, neoxam_green),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 30))

            data_title = Paragraph("📋 Données Extraites", subtitle_style)
            elements.append(data_title)

            note_text = f"""
            <i>Le tableau ci-dessous présente les {total_options} options analysées avec leurs variables et valeurs correspondantes.</i>
            """
            note = Paragraph(note_text, normal_style)
            elements.append(note)
            elements.append(Spacer(1, 15))

            if type_etat == 1:
                headers = ['🔧 Option', '📊 Variable', '💡 Valeur']
                col_widths = [2*inch, 2.5*inch, 2.5*inch]
                table_data = [headers]
                for option in self.parsed_data.get('options', []):
                    option_name = option.get('name', '')
                    for variable in option.get('variables', []):
                        opt = Paragraph(option_name or '', table_cell_style)
                        var = Paragraph(variable.get('name', '') or '', table_cell_style)
                        val = Paragraph(variable.get('value', '') or '', table_cell_style)
                        table_data.append([opt, var, val])
                        option_name = ''
                    if not option.get('variables', []):
                        table_data.append([Paragraph(option_name or '', table_cell_style), '', ''])
            else:
                headers = ['🔧 Option', '⚙️ Procédure', '📊 Variable', '💡 Valeur']
                col_widths = [1.5*inch, 2*inch, 2*inch, 2*inch]
                table_data = [headers]
                for option in self.parsed_data.get('options', []):
                    option_name = option.get('name', '')
                    procedure = option.get('procedure', '')
                    for variable in option.get('variables', []):
                        opt = Paragraph(option_name or '', table_cell_style)
                        proc = Paragraph(procedure or '', table_cell_style)
                        var = Paragraph(variable.get('name', '') or '', table_cell_style)
                        val = Paragraph(variable.get('value', '') or '', table_cell_style)
                        table_data.append([opt, proc, var, val])
                        option_name = ''
                        procedure = ''
                    if not option.get('variables', []):
                        table_data.append([Paragraph(option_name or '', table_cell_style),
                                         Paragraph(procedure or '', table_cell_style), '', ''])

            main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            main_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), neoxam_green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 1), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, neoxam_light_gray]),
                ('LINEBELOW', (0, 0), (-1, 0), 2, neoxam_green),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6')),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8F5E8')),
            ]))
            elements.append(main_table)
            elements.append(Spacer(1, 30))

            stats_title = Paragraph("📈 Statistiques", subtitle_style)
            elements.append(stats_title)

            stats_text = f"""
            <b>Analyse complète :</b><br/>
            • <b>{total_options}</b> options analysées<br/>
            • <b>{total_variables}</b> variables extraites<br/>
            • <b>{unique_options}</b> options uniques<br/>
            • Fichier source : <i>{os.path.basename(self.parsed_data.get('filepath', 'Unknown'))}</i>
            """
            stats_para = Paragraph(stats_text, normal_style)
            elements.append(stats_para)

            elements.append(Spacer(1, 20))
            footer_text = f"""
            <b>🔒 Rapport Confidentiel NEOXAM</b><br/>
            <i>Ce document a été généré automatiquement par l'application XML Parser de NEOXAM.<br/>
            Toute reproduction ou diffusion est soumise à autorisation.</i>
            """
            footer_style = ParagraphStyle(
                'FooterStyle',
                parent=normal_style,
                fontSize=9,
                textColor=neoxam_gray,
                alignment=TA_CENTER,
                borderColor=neoxam_green,
                borderWidth=1,
                borderPadding=10
            )
            footer = Paragraph(footer_text, footer_style)
            elements.append(footer)

            doc.build(elements)
            if self.log_func:
                self.log_func(f"Generated PDF: {pdf_path}")
            return pdf_path
        except Exception as e:
            if self.log_func:
                self.log_func(f"Error generating PDF: {e}")
            print(f"Error generating PDF: {e}")
            return None

    def generate_csv(self, csv_path):
        """Generate CSV report with option, procedure, variable, and value"""
        try:
            os.makedirs(os.path.dirname(csv_path) or "csvs", exist_ok=True)
            data = [['Option', 'Procédure', 'Variable', 'Valeur']]
            for option in self.parsed_data.get('options', []):
                option_name = option.get('name', '')
                procedure = option.get('procedure', '')
                for variable in option.get('variables', []):
                    data.append([option_name, procedure, variable.get('name', ''), variable.get('value', '')])
                    option_name = ''
                    procedure = ''
                if not option.get('variables', []):
                    data.append([option_name, procedure, '', ''])

            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                for row in data:
                    writer.writerow(row)

            if self.log_func:
                self.log_func(f"Generated CSV: {csv_path}")
            return csv_path
        except Exception as e:
            if self.log_func:
                self.log_func(f"Error generating CSV: {e}")
            print(f"Error generating CSV: {e}")
            return None

    def generate_excel(self, excel_path):
        """Generate Excel report with two sheets: full data and unique options with occurrence counts"""
        try:
            if self.log_func:
                self.log_func(f"Starting Excel generation for: {excel_path}")
                self.log_func(f"Parsed data keys: {list(self.parsed_data.keys())}")
                self.log_func(f"Number of options: {len(self.parsed_data.get('options', []))}")

            os.makedirs(os.path.dirname(excel_path) or "reports", exist_ok=True)
            wb = openpyxl.Workbook()
            
            # Main sheet: Full data
            if self.log_func:
                self.log_func("Creating main sheet 'Rapport XML'")
            ws_main = wb.active
            ws_main.title = "Rapport XML"
            headers = ['Option', 'Procédure', 'Variable', 'Valeur']
            ws_main.append(headers)

            options = self.parsed_data.get('options', [])
            if self.log_func:
                self.log_func(f"Processing {len(options)} options for main sheet")
            for option in options:
                option_name = str(option.get('name', '')) if option.get('name') is not None else ''
                procedure = str(option.get('procedure', '')) if option.get('procedure') is not None else ''
                variables = option.get('variables', [])
                if self.log_func:
                    self.log_func(f"Processing option: {option_name}, procedure: {procedure}, variables: {len(variables)}")
                if variables:
                    for variable in variables:
                        ws_main.append([
                            option_name,
                            procedure,
                            str(variable.get('name', '')) if variable.get('name') is not None else '',
                            str(variable.get('value', '')) if variable.get('value') is not None else ''
                        ])
                        option_name = ''
                        procedure = ''
                else:
                    ws_main.append([option_name, procedure, '', ''])

            # Apply styling to main sheet headers
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="00C4B4", end_color="00C4B4", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws_main[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Adjust column widths for main sheet
            for col in ws_main.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_main.column_dimensions[column].width = adjusted_width

            ws_main.auto_filter.ref = ws_main.dimensions
            if self.log_func:
                self.log_func("Main sheet 'Rapport XML' created successfully")

            # Unique Options sheet with occurrence counts
            if self.log_func:
                self.log_func("Creating 'Options Uniques' sheet")
            ws_options = wb.create_sheet(title="Options Uniques")
            ws_options.append(['Option Unique', 'Occurrences'])
            
            # Collect unique options and their counts
            option_counts = Counter(
                str(option.get('name', '')) for option in self.parsed_data.get('options', [])
                if option.get('name') is not None and str(option.get('name')).strip()
            )
            
            if self.log_func:
                self.log_func(f"Found {len(option_counts)} unique options: {list(option_counts.keys())}")
            
            if not option_counts:
                if self.log_func:
                    self.log_func("No unique options found, adding placeholder")
                ws_options.append(['Aucune option unique trouvée', 0])
                max_option_length = 20  # Default width for "Aucune option unique trouvée"
            else:
                for option_name, count in sorted(option_counts.items()):
                    ws_options.append([option_name, count])
                if self.log_func:
                    self.log_func(f"Added {len(option_counts)} unique options with counts to sheet")
                max_option_length = max(len(str(option)) for option in option_counts.keys()) if option_counts else 15

            # Apply styling to unique options sheet headers
            for cell in ws_options[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Adjust column widths for unique options sheet
            ws_options.column_dimensions['A'].width = min(max_option_length + 2, 50)
            ws_options.column_dimensions['B'].width = 15  # Fixed width for occurrence count

            ws_options.auto_filter.ref = ws_options.dimensions
            if self.log_func:
                self.log_func("Unique options sheet created successfully")

            # Verify sheets in workbook
            sheet_names = wb.sheetnames
            if self.log_func:
                self.log_func(f"Sheets in workbook: {sheet_names}")
            if "Options Uniques" not in sheet_names:
                if self.log_func:
                    self.log_func("Error: 'Options Uniques' sheet not found in workbook")

            # Save the workbook
            if self.log_func:
                self.log_func(f"Saving Excel file to {excel_path}")
            wb.save(excel_path)
            if self.log_func:
                self.log_func(f"Generated Excel with sheets: {sheet_names}")
            return excel_path
        except Exception as e:
            if self.log_func:
                self.log_func(f"Error generating Excel: {str(e)}")
            print(f"Error generating Excel: {str(e)}")
            return None

class LoginWindow(QWidget):
    def __init__(self):
        self.db_manager = DatabaseManager()
        self.current_user = None
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Authentification - NeoXAm')
        self.setWindowIcon(QIcon("images/iconapp.ico"))
        self.setFixedSize(850, 850)
        self.setStyleSheet("""
             QWidget {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #000000,
                    stop: 0.4 #1a1a2e,
                    stop: 1 #0f2f2f
                );
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                color: #FFFFFF;
                font-size: 16px;
                font-weight: bold;
                background: transparent;
            }
            QLineEdit {
                background-color: #1C2526;
                color: #FFFFFF;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 12px;
                padding: 12px;
                font-size: 14px;
                selection-background-color: #12e0d6;
            }
            QLineEdit:focus {
                border: 2px solid #0fb8b0;
                outline: none;
                background-color: #1C2526;
            }
            QPushButton {
                background-color: #12e0d6;
                color: #000000;
                border: 2px solid #0fb8b0;
                border-radius: 12px;
                padding: 12px 20px;
                font-size: 16px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0fb8b0;
                border-color: #0d8f88;
                transform: translateY(-3px);
                transition: all 0.3s ease-in-out;
            }
            QPushButton:pressed {
                background-color: #0d8f88;
                transform: translateY(0);
            }
            QHBoxLayout {
                justify-content: center;
                spacing: 15px;
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(25)
        layout.setContentsMargins(50, 50, 50, 50)

        self.logo_label = QLabel(self)
        pixmap = QtGui.QPixmap('images/logo.png').scaled(300, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        if pixmap.isNull():
            self.logo_label.setText("Logo non trouvé (images/logo.png)")
            self.logo_label.setStyleSheet("color: #22d3ee; font-size: 14px;")
        else:
            self.logo_label.setPixmap(pixmap)
        self.logo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.logo_label)

        current_time = datetime.now().strftime("%I:%M %p CET, %d %B %Y")
        welcome_msg = QLabel(f"Bienvenue ! Il est {current_time}", self)
        welcome_msg.setAlignment(Qt.AlignCenter)
        welcome_msg.setStyleSheet("font-size: 14px; color: #22d3ee;")
        layout.addWidget(welcome_msg)

        title = QLabel('Authentification', self)
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 26px; color: #22d3ee;")
        layout.addWidget(title)

        self.username_label = QLabel('Nom d\'utilisateur:', self)
        self.username_input = QLineEdit(self)
        self.username_input.setPlaceholderText('Entrez votre nom d\'utilisateur')
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)

        self.password_layout = QHBoxLayout()
        self.password_label = QLabel('Mot de passe:', self)
        self.password_input = QLineEdit(self)
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText('Entrez votre mot de passe')
        self.toggle_password_button = QPushButton('👁️', self)
        self.toggle_password_button.setFixedSize(40, 40)
        self.toggle_password_button.setStyleSheet("""
            background-color: #1e293b;
            border: none;
            color: #22d3ee;
            font-size: 16px;
            padding: 0;
        """)
        self.toggle_password_button.clicked.connect(self.toggle_password_visibility)
        self.password_layout.addWidget(self.password_input)
        self.password_layout.addWidget(self.toggle_password_button)
        layout.addWidget(self.password_label)
        layout.addLayout(self.password_layout)

        button_layout = QHBoxLayout()
        self.login_button = QPushButton('Connexion', self)
        self.login_button.clicked.connect(self.check_credentials)
        self.quit_button = QPushButton('Quitter', self)
        self.quit_button.clicked.connect(self.close)
        self.reset_button = QPushButton('Réinitialiser', self)
        self.reset_button.clicked.connect(self.reset_fields)
        button_layout.addWidget(self.login_button)
        button_layout.addWidget(self.quit_button)
        button_layout.addWidget(self.reset_button)
        layout.addLayout(button_layout)

         
        self.register_button = QPushButton('S\'inscrire')
        self.register_button.clicked.connect(self.show_register)
        button_layout.addWidget(self.register_button)

        forgot_pwd = QLabel('<a href="#" style="color: #22d3ee; text-decoration: none;">Mot de passe oublié ?</a>', self)
        forgot_pwd.setAlignment(Qt.AlignCenter)
        forgot_pwd.setOpenExternalLinks(False)
        forgot_pwd.mousePressEvent = lambda x: self.show_forgot_pwd_dialog()
        layout.addWidget(forgot_pwd)

        self.setLayout(layout)

    def toggle_password_visibility(self):
        if self.password_input.echoMode() == QLineEdit.Password:
            self.password_input.setEchoMode(QLineEdit.Normal)
            self.toggle_password_button.setText('🔒')
        else:
            self.password_input.setEchoMode(QLineEdit.Password)
            self.toggle_password_button.setText('👁️')

    def check_credentials(self):
        username = self.username_input.text().strip()
        password = self.password_input.text()

        if not username or not password:
            QMessageBox.warning(self, 'Erreur', 'Veuillez remplir tous les champs.')
            return

        success, result = self.db_manager.authenticate_user(username, password)

        if success:
            self.current_user = result
            QMessageBox.information(self, 'Succès', 'Connexion réussie!')
            self.close()
            
            # Ouvrir l'interface admin si c'est un admin
            if result['role'] == 'admin':
                self.show_admin_or_main()
            else:
                self.main_window = XmlParserApp(self.current_user)
                self.main_window.show()
        else:
            QMessageBox.warning(self, 'Erreur', result)

    def show_admin_or_main(self):
        reply = QMessageBox.question(self, 'Choix d\'interface', 
                                   'Voulez-vous accéder à l\'interface d\'administration?',
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.admin_window = AdminWindow(self.current_user)
            self.admin_window.show()
        else:
            self.main_window = XmlParserApp(self.current_user)
            self.main_window.show()

    def show_register(self):
        self.register_window = RegisterWindow()
        self.register_window.show()


    def reset_fields(self):
        self.username_input.clear()
        self.password_input.clear()

    def show_forgot_pwd_dialog(self, link):
        if link == "forgot-password":
            QMessageBox.information(self, 'Mot de passe oublié', 'Veuillez contacter l\'administrateur pour réinitialiser votre mot de passe.', QMessageBox.Ok)

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.page_count = 0

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()
        self.page_count += 1

    def save(self):
        num_pages = len(self._saved_page_states)
        for (page_num, page_state) in enumerate(self._saved_page_states):
            self.__dict__.update(page_state)
            self.draw_page_elements(page_num + 1, num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_elements(self, page_num, total_pages):
        self.setStrokeColor(colors.HexColor('#00C4B4'))
        self.setLineWidth(3)
        self.line(50, A4[1] - 70, A4[0] - 50, A4[1] - 70)
        
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor('#4B5563'))
        self.drawRightString(A4[0] - 50, 30, f"Page {page_num} sur {total_pages}")
        self.drawString(50, 30, f"NEOXAM - Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")

def create_neoxam_logo_placeholder():
    try:
        logo_path = "images/pdflogo.png"
        if os.path.exists(logo_path):
            return Image(logo_path, width=2.5*inch, height=0.8*inch)
        else:
            return None
    except:
        return None
    




class DashboardWindow(QDialog):
    def __init__(self, csv_data, qr_path, parent=None):
        super().__init__(parent)
        self.csv_data = csv_data or [] # Ensure csv_data is not None
        self.qr_path = qr_path
        # Debug: Log raw csv_data and export to CSV
        print(f"[DEBUG] DashboardWindow initialized with {len(self.csv_data)} rows in csv_data")
        self.export_csv_data() # Export csv_data for inspection
        self.initUI()
        # Debug: Log computed option counts
        option_counts = self.get_option_counts()
        print(f"[DEBUG] Computed {len(option_counts)} unique options: {list(option_counts.items())[:10]}")

    def export_csv_data(self):
        """Export csv_data to a CSV file for debugging."""
        try:
            with open('debug_csv_data.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(['Option', 'Procedure', 'Variable', 'Value'])
                writer.writerows(self.csv_data)
            print("[DEBUG] Exported csv_data to debug_csv_data.csv")
        except Exception as e:
            print(f"[ERROR] Failed to export csv_data: {e}")

    def get_option_counts(self):
        """Count occurrences of each unique option based solely on non-empty values in the 'Option' column (row[0]).
        This ensures we only count explicit option declarations, ignoring any potential matches in procedures, variables, or values.
        Each non-empty option_name increments the count for that option, handling potential duplicates."""
        option_counts = Counter()
        for row in self.csv_data:
            if len(row) < 4:
                continue  # Skip invalid rows
            option_name = row[0]
            if option_name and isinstance(option_name, str) and option_name.strip():
                option_counts[option_name.strip()] += 1
        if not option_counts:
            print("[WARNING] No valid options found in csv_data")
        return option_counts

    def get_filtered_field(self, column):
        """Extract valid values from a specified column in csv_data."""
        return [row[column] for row in self.csv_data if len(row) > column and row[column] and isinstance(row[column], str) and row[column].strip()]

    def initUI(self):
        self.setWindowTitle('🚀 Dashboard XML - Analyse Avancée')
        self.setWindowIcon(QIcon("images/iconapp.ico"))
        self.setFixedSize(1600, 1200)
        self.setStyleSheet("""
            QWidget {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #000000,
                    stop: 0.4 #1a1a2e,
                    stop: 1 #0f2f2f
                );
                font-family: 'Segoe UI', Arial, sans-serif;
                color: #FFFFFF;
            }
            QLabel {
                color: #FFFFFF;
                background: transparent;
            }
            QPushButton {
                background-color: #12e0d6;
                color: #000000;
                border: 2px solid #0fb8b0;
                border-radius: 12px;
                font-weight: bold;
                font-size: 14px;
                padding: 12px 20px;
                min-height: 50px;
            }
            QPushButton:hover {
                background-color: #0fb8b0;
                border: 2px solid #0d8f88;
                transform: translateY(-3px);
            }
            QPushButton:pressed {
                background-color: #0d8f88;
                border: 2px solid #0d8f88;
                transform: translateY(0px);
            }
            QTableWidget {
                background-color: #1C2526;
                color: #FFFFFF;
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 12px;
                font-size: 14px;
                gridline-color: #0f2f2f;
                selection-background-color: #12e0d6;
                selection-color: #000000;
            }
            QTableWidget::item {
                padding: 14px 10px;
                border: none;
            }
            QHeaderView::section {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                color: #000000;
                font-weight: bold;
                font-size: 15px;
                padding: 18px 10px;
                border: none;
                border-bottom: 4px solid #0d8f88;
            }
            QFrame#cardFrame {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #000000,
                    stop: 1 #1a1a2e
                );
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 12px;
                padding: 25px;
            }
            QFrame#statCard {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 rgba(18, 224, 214, 0.15),
                    stop: 1 rgba(15, 184, 176, 0.15)
                );
                border: 2px solid #12e0d6;
                border-radius: 12px;
                padding: 20px;
                min-height: 140px;
            }
            QFrame#statCard:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 rgba(18, 224, 214, 0.25),
                    stop: 1 rgba(15, 184, 176, 0.25)
                );
                border: 2px solid #0fb8b0;
                transform: scale(1.02);
            }
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                background: rgba(15, 47, 47, 0.3);
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #12e0d6;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #0fb8b0;
            }
        """)
        # Main scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
       
        main_widget = QWidget()
        self.main_layout = QVBoxLayout(main_widget)
        self.main_layout.setSpacing(30)
        self.main_layout.setContentsMargins(40, 40, 40, 40)
        # Header Section
        self.create_header()
       
        # Large Statistics Section
        self.create_large_stats_section()
       
        # Chart Controls Section
        self.create_chart_controls()
       
        # Chart Section
        self.create_chart_section()
       
        # Options Table Section
        self.create_options_table_section()
       
        # QR Code Section
        self.create_qr_section()
       
        # Export Section
        self.create_export_section()
        scroll.setWidget(main_widget)
       
        main_window_layout = QVBoxLayout()
        main_window_layout.setContentsMargins(0, 0, 0, 0)
        main_window_layout.addWidget(scroll)
        self.setLayout(main_window_layout)
       
        self.show_bar_chart()

    def create_header(self):
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 rgba(18, 224, 214, 0.2),
                    stop: 0.5 rgba(15, 184, 176, 0.2),
                    stop: 1 rgba(13, 143, 136, 0.2)
                );
                border: 2px solid #12e0d6;
                border-radius: 12px;
                padding: 35px;
            }
        """)
        header_layout = QVBoxLayout(header_frame)
       
        title_label = QLabel('🚀 DASHBOARD XML - ANALYSE COMPLÈTE')
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 38px;
            font-weight: 900;
            color: #12e0d6;
            margin-bottom: 15px;
            background: transparent;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5);
        """)
       
        subtitle_label = QLabel('Visualisation avancée et statistiques détaillées de vos données XML')
        subtitle_label.setAlignment(Qt.AlignCenter)
        subtitle_label.setStyleSheet("""
            font-size: 18px;
            color: #FFFFFF;
            font-style: italic;
            font-weight: 500;
            background: transparent;
        """)
       
        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        self.main_layout.addWidget(header_frame)

    def create_large_stats_section(self):
        stats_frame = QFrame()
        stats_frame.setObjectName('cardFrame')
        stats_layout = QVBoxLayout(stats_frame)
       
        stats_title = QLabel('📊 STATISTIQUES PRINCIPALES')
        stats_title.setStyleSheet("""
            font-size: 24px;
            font-weight: 900;
            color: #12e0d6;
            margin-bottom: 25px;
            background: transparent;
        """)
        stats_layout.addWidget(stats_title)
       
        # Calculate comprehensive statistics
        options_data = self.csv_data
        options = [row[0] for row in options_data if row[0] and row[0].strip()]
        nb_option = len(options)
        unique_options = len(set(options))
        most_common_option = Counter(options).most_common(1)[0] if options else ("N/A", 0)
       
        procedures = [row[1] for row in options_data if row[1] and row[1].strip()]
        unique_procedures = len(set(procedures))
       
        variables = [row[2] for row in options_data if row[2] and row[2].strip()]
        unique_variables = len(set(variables))
       
        values = [row[3] for row in options_data if row[3] and row[3].strip()]
        avg_value_length = np.mean([len(v) for v in values]) if values else 0
       
        most_common_procedure = Counter(procedures).most_common(1)[0] if procedures else ("N/A", 0)
       
        empty_values = len([row for row in options_data if not row[0] and not row[1] and not row[2] and not row[3]])
        completion_rate = ((len(options_data) - empty_values) / len(options_data) * 100) if len(options_data) > 0 else 0
       
        # Create large stats grid
        stats_grid = QGridLayout()
        stats_grid.setSpacing(25)
       
        # Row 1 - Main metrics
        stats_grid.addWidget(self.create_stat_card("OPTIONS UNIQUES", unique_options, "Types d'options différents", "🎯", "#12e0d6", True), 0, 0)
        stats_grid.addWidget(self.create_stat_card("PROCÉDURES", unique_procedures, "Méthodes distinctes", "⚙️", "#0fb8b0", True), 0, 1)
        stats_grid.addWidget(self.create_stat_card("VARIABLES", unique_variables, "Paramètres uniques", "📋", "#0d8f88", True), 0, 2)
       
        # Row 2 - Secondary metrics
        stats_grid.addWidget(self.create_stat_card("TOTAL ENTRÉES", len(options_data), "Lignes de données", "📊", "#12e0d6", True), 1, 0)
        stats_grid.addWidget(self.create_stat_card("LONGUEUR MOY.", f"{avg_value_length:.1f}", "Caractères par valeur", "📏", "#0fb8b0", True), 1, 1)
        stats_grid.addWidget(self.create_stat_card("TAUX COMPLÉTUDE", f"{completion_rate:.1f}%", "Données complètes", "✅", "#0d8f88", True), 1, 2)
       
        # Row 3 - Popular items
        stats_grid.addWidget(self.create_stat_card("OPTION POPULAIRE", most_common_option[0], f"{most_common_option[1]} occurrences", "🏆", "#12e0d6"), 2, 0, 1, 2)
        stats_grid.addWidget(self.create_stat_card("PROCÉDURE POPULAIRE", most_common_procedure[0], f"{most_common_procedure[1]} occurrences", "⭐", "#0fb8b0"), 2, 2)
       
        stats_layout.addLayout(stats_grid)
        self.main_layout.addWidget(stats_frame)

    def create_stat_card(self, title, value, subtitle, icon, color, large_number=False):
        card = QFrame()
        card.setObjectName('statCard')
        card.setStyleSheet(f"""
            QFrame#statCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {color}15,
                    stop: 1 {color}25
                );
                border: 2px solid {color};
                border-radius: 12px;
                padding: 25px;
                min-height: 160px;
            }}
            QFrame#statCard:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {color}25,
                    stop: 1 {color}35
                );
                border: 2px solid {color};
                transform: scale(1.02);
            }}
        """)
       
        layout = QVBoxLayout(card)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)
       
        # Icon and value row
        top_layout = QHBoxLayout()
       
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"""
            font-size: 45px;
            color: {color};
            background: transparent;
        """)
        icon_label.setAlignment(Qt.AlignLeft)
       
        value_label = QLabel(str(value))
        if large_number:
            value_label.setStyleSheet(f"""
                font-size: 42px;
                font-weight: 900;
                color: {color};
                background: transparent;
            """)
        else:
            value_label.setStyleSheet(f"""
                font-size: 32px;
                font-weight: 800;
                color: {color};
                background: transparent;
            """)
        value_label.setAlignment(Qt.AlignRight)
       
        top_layout.addWidget(icon_label)
        top_layout.addStretch()
        top_layout.addWidget(value_label)
       
        # Title
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 700;
            color: #FFFFFF;
            background: transparent;
        """)
        title_label.setAlignment(Qt.AlignLeft)
       
        # Subtitle
        subtitle_label = QLabel(subtitle)
        subtitle_label.setStyleSheet("""
            font-size: 13px;
            color: #FFFFFF;
            background: transparent;
            font-weight: 500;
        """)
        subtitle_label.setAlignment(Qt.AlignLeft)
       
        layout.addLayout(top_layout)
        layout.addWidget(title_label)
        layout.addWidget(subtitle_label)
        layout.addStretch()
       
        return card

    def create_chart_controls(self):
        controls_frame = QFrame()
        controls_frame.setObjectName('cardFrame')
        controls_layout = QVBoxLayout(controls_frame)
       
        controls_title = QLabel('📈 SÉLECTION DE GRAPHIQUE')
        controls_title.setStyleSheet("""
            font-size: 22px;
            font-weight: 900;
            color: #12e0d6;
            margin-bottom: 20px;
            background: transparent;
        """)
        controls_layout.addWidget(controls_title)
       
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
       
        self.btn_bar = QPushButton('📊 Graphique en Barres')
        self.btn_bar.clicked.connect(self.show_bar_chart)
        self.btn_pie = QPushButton('🥧 Graphique Circulaire')
        self.btn_pie.clicked.connect(self.show_pie_chart)
        self.btn_wordcloud = QPushButton('☁️ Nuage de Mots')
        self.btn_wordcloud.clicked.connect(self.show_word_cloud)
        self.btn_histogram = QPushButton('📈 Histogramme')
        self.btn_histogram.clicked.connect(self.show_histogram)
        self.btn_boxplot = QPushButton('📉 Boîte à Moustaches')
        self.btn_boxplot.clicked.connect(self.show_boxplot)
       
        for btn in [self.btn_bar, self.btn_pie, self.btn_wordcloud, self.btn_histogram, self.btn_boxplot]:
            button_layout.addWidget(btn)
       
        controls_layout.addLayout(button_layout)
        self.main_layout.addWidget(controls_frame)

    def create_chart_section(self):
        chart_frame = QFrame()
        chart_frame.setObjectName('cardFrame')
        chart_layout = QVBoxLayout(chart_frame)
       
        chart_title = QLabel('📊 VISUALISATION DES DONNÉES')
        chart_title.setStyleSheet("""
            font-size: 22px;
            font-weight: 900;
            color: #12e0d6;
            margin-bottom: 20px;
            background: transparent;
        """)
        chart_layout.addWidget(chart_title)
       
        self.figure, self.ax = plt.subplots(figsize=(14, 8))
        self.canvas = FigureCanvas(self.figure)
        chart_layout.addWidget(self.canvas)
        self.main_layout.addWidget(chart_frame, stretch=2)

    def create_options_table_section(self):
        table_frame = QFrame()
        table_frame.setObjectName('cardFrame')
        table_layout = QVBoxLayout(table_frame)
       
        table_title = QLabel('📋 OPTIONS XML DÉTECTÉES (TOP 100)')
        table_title.setStyleSheet("""
            font-size: 22px;
            font-weight: 900;
            color: #12e0d6;
            margin-bottom: 20px;
            background: transparent;
        """)
        table_layout.addWidget(table_title)
       
        # Use option counts
        option_counts = self.get_option_counts()
        sorted_options = option_counts.most_common(100) # Top 100 options
       
        self.data_table = QTableWidget()
        self.data_table.setRowCount(len(sorted_options))
        self.data_table.setColumnCount(4)
        self.data_table.setHorizontalHeaderLabels(['#', 'Option XML', 'Occurrences', 'Pourcentage'])
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setStyleSheet("""
            QTableWidget::item:alternate {
                background-color: rgba(18, 224, 214, 0.1);
            }
            QTableWidget::item {
                background-color: #1C2526;
            }
        """)
       
        total_options = sum(option_counts.values()) if option_counts else 0
        if not sorted_options:
            print("[WARNING] No options to display in options table")
            self.data_table.setRowCount(1)
            self.data_table.setItem(0, 0, QTableWidgetItem("N/A"))
            self.data_table.setItem(0, 1, QTableWidgetItem("Aucune option disponible"))
            self.data_table.setItem(0, 2, QTableWidgetItem("0"))
            self.data_table.setItem(0, 3, QTableWidgetItem("0.00%"))
        else:
            for i, (option, count) in enumerate(sorted_options):
                percentage = (count / total_options) * 100 if total_options > 0 else 0
               
                rank_item = QTableWidgetItem(str(i + 1))
                rank_item.setTextAlignment(Qt.AlignCenter)
                rank_item.setFont(QFont("Arial", 12, QFont.Bold))
               
                option_item = QTableWidgetItem(option)
                option_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                option_item.setFont(QFont("Consolas", 11))
               
                count_item = QTableWidgetItem(str(count))
                count_item.setTextAlignment(Qt.AlignCenter)
                count_item.setFont(QFont("Arial", 12, QFont.Bold))
               
                percentage_item = QTableWidgetItem(f"{percentage:.2f}%")
                percentage_item.setTextAlignment(Qt.AlignCenter)
                percentage_item.setFont(QFont("Arial", 11))
               
                # Color coding based on frequency
                if percentage >= 10:
                    color = QColor(18, 224, 214, 100) # High frequency
                elif percentage >= 5:
                    color = QColor(15, 184, 176, 80) # Medium frequency
                else:
                    color = QColor(13, 143, 136, 60) # Low frequency
               
                for item in [rank_item, option_item, count_item, percentage_item]:
                    item.setBackground(color)
               
                self.data_table.setItem(i, 0, rank_item)
                self.data_table.setItem(i, 1, option_item)
                self.data_table.setItem(i, 2, count_item)
                self.data_table.setItem(i, 3, percentage_item)
       
        self.data_table.setFixedHeight(450)
        self.data_table.resizeColumnsToContents()
        self.data_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table_layout.addWidget(self.data_table)
        self.main_layout.addWidget(table_frame)

    def create_qr_section(self):
        qr_frame = QFrame()
        qr_frame.setObjectName('cardFrame')
        qr_layout = QHBoxLayout(qr_frame)
       
        qr_info_layout = QVBoxLayout()
        qr_title = QLabel('🔗 QR CODE DU CONTENU XML')
        qr_title.setStyleSheet("""
            font-size: 22px;
            font-weight: 900;
            color: #12e0d6;
            background: transparent;
        """)
        qr_info = QLabel('Scannez ce code QR pour accéder rapidement au contenu XML original.\nParfait pour partager ou accéder aux données depuis un appareil mobile.')
        qr_info.setStyleSheet("""
            font-size: 16px;
            color: #FFFFFF;
            background: transparent;
            line-height: 1.5;
        """)
        qr_info.setWordWrap(True)
       
        qr_info_layout.addWidget(qr_title)
        qr_info_layout.addWidget(qr_info)
        qr_info_layout.addStretch()
       
        self.qr_label = QLabel()
        if self.qr_path and os.path.exists(self.qr_path):
            pixmap = QPixmap(self.qr_path).scaled(220, 220, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.qr_label.setPixmap(pixmap)
            self.qr_label.setStyleSheet("""
                border: 2px solid #12e0d6;
                border-radius: 12px;
                padding: 15px;
                background: rgba(18, 224, 214, 0.1);
            """)
        else:
            self.qr_label.setText("❌ QR CODE\nNON DISPONIBLE")
            self.qr_label.setStyleSheet("""
                color: #f56565;
                font-size: 16px;
                font-weight: bold;
                text-align: center;
                border: 2px dashed #f56565;
                border-radius: 12px;
                padding: 30px;
                background: rgba(245, 101, 101, 0.1);
            """)
       
        self.qr_label.setAlignment(Qt.AlignCenter)
        self.qr_label.setFixedSize(250, 250)
       
        qr_layout.addLayout(qr_info_layout)
        qr_layout.addStretch()
        qr_layout.addWidget(self.qr_label)
        self.main_layout.addWidget(qr_frame)

    def create_export_section(self):
        export_frame = QFrame()
        export_frame.setObjectName('cardFrame')
        export_layout = QHBoxLayout(export_frame)
       
        export_info = QLabel('💾 Exportez vos graphiques en haute qualité (PNG/PDF/SVG)')
        export_info.setStyleSheet("""
            font-size: 18px;
            font-weight: 700;
            color: #FFFFFF;
            background: transparent;
        """)
       
        self.btn_export = QPushButton('💾 EXPORTER LE GRAPHIQUE ACTUEL')
        self.btn_export.setFixedHeight(60)
        self.btn_export.clicked.connect(self.export_chart)
        self.btn_export.setStyleSheet("""
            QPushButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                color: #000000;
                font-size: 16px;
                font-weight: 900;
                min-width: 280px;
                border-radius: 12px;
                border: 2px solid #0fb8b0;
            }
            QPushButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #0fb8b0,
                    stop: 1 #0d8f88
                );
                border: 2px solid #0d8f88;
                box-shadow: 0 10px 25px rgba(18, 224, 214, 0.4);
            }
        """)
       
        export_layout.addWidget(export_info)
        export_layout.addStretch()
        export_layout.addWidget(self.btn_export)
        self.main_layout.addWidget(export_frame)

    def clear_axes(self):
        self.ax.clear()
        self.figure.patch.set_facecolor('#1a1a2e')
        self.ax.set_facecolor('#1a1a2e')
        self.ax.tick_params(colors='#FFFFFF', labelsize=12)
        self.ax.title.set_color('#12e0d6')
        self.ax.xaxis.label.set_color('#FFFFFF')
        self.ax.yaxis.label.set_color('#FFFFFF')
        for spine in self.ax.spines.values():
            spine.set_color('#0f2f2f')
            spine.set_linewidth(2)

    def show_bar_chart(self):
        self.clear_axes()
        option_counts = self.get_option_counts()
        if not option_counts:
            self.ax.text(0.5, 0.5, 'AUCUNE OPTION DISPONIBLE',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=20, color='#f56565', weight='bold')
            self.canvas.draw()
            return
       
        top_options = option_counts.most_common(20) # Top 20 options
       
        labels = [opt[0] for opt in top_options]
        values = [opt[1] for opt in top_options]
       
        # Create gradient colors
        colors = ['#12e0d6', '#0fb8b0', '#0d8f88', '#15c7c0', '#18d0ca'] * 4
       
        bars = self.ax.bar(labels, values,
                          color=colors[:len(labels)],
                          edgecolor='#0f2f2f',
                          linewidth=2,
                          alpha=0.9)
       
        self.ax.set_title('TOP 20 DES OPTIONS XML LES PLUS FRÉQUENTES', fontsize=18, pad=25, weight='bold')
        self.ax.set_xlabel('Options XML', fontsize=14, weight='700')
        self.ax.set_ylabel('Nombre d\'Occurrences', fontsize=14, weight='700')
        self.ax.tick_params(axis='x', rotation=45, labelsize=11)
        self.ax.grid(True, axis='y', linestyle='--', alpha=0.5, color='#FFFFFF')
       
        # Add value labels on bars
        for bar, value in zip(bars, values):
            height = bar.get_height()
            self.ax.text(bar.get_x() + bar.get_width()/2., height + max(values)*0.01,
                        f'{value}', ha='center', va='bottom', fontsize=11,
                        color='#FFFFFF', weight='bold')
       
        mplcursors.cursor(bars).connect("add", lambda sel: sel.annotation.set_text(f"{labels[sel.index]}: {values[sel.index]} occurrences"))
        self.figure.tight_layout()
        self.canvas.draw()

    def show_pie_chart(self):
        self.clear_axes()
        option_counts = self.get_option_counts()
        if not option_counts:
            self.ax.text(0.5, 0.5, 'AUCUNE OPTION DISPONIBLE',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=20, color='#f56565', weight='bold')
            self.canvas.draw()
            return
       
        top_options = option_counts.most_common(10) # Top 10 for better visibility
       
        labels = [opt[0][:25] + "..." if len(opt[0]) > 25 else opt[0] for opt in top_options]
        values = [opt[1] for opt in top_options]
       
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', 
 '#FF9FF3', '#54A0FF', '#5F27CD', "#EEFF00", "#49FF43"]
       
        wedges, texts, autotexts = self.ax.pie(values, labels=labels, colors=colors,
                                              autopct='%1.2f%%', startangle=90,
                                              textprops={'fontsize': 12, 'weight': 'bold'})
       
        for text in texts + autotexts:
            text.set_color('#FFFFFF')
       
        self.ax.set_title('RÉPARTITION DES OPTIONS XML PRINCIPALES', fontsize=18, pad=25, weight='bold')
        mplcursors.cursor(wedges).connect("add", lambda sel: sel.annotation.set_text(f"{labels[sel.index]}: {values[sel.index]} ({autotexts[sel.index].get_text()})"))
        self.figure.tight_layout()
        self.canvas.draw()

    def show_word_cloud(self):
        self.clear_axes()
        option_counts = self.get_option_counts()
        if not option_counts:
            self.ax.text(0.5, 0.5, 'AUCUNE OPTION DISPONIBLE',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=20, color='#f56565', weight='bold')
            self.canvas.draw()
            return
       
        wordcloud = WordCloud(
            width=1400, height=700,
            background_color='#1a1a2e',
            colormap='Blues',
            min_font_size=16,
            max_font_size=100,
            relative_scaling=0.7,
            max_words=150,
            prefer_horizontal=0.7
        ).generate_from_frequencies(option_counts)
       
        self.ax.imshow(wordcloud, interpolation='bilinear')
        self.ax.axis('off')
        self.ax.set_title('NUAGE DE MOTS DES OPTIONS XML', fontsize=18, pad=25, weight='bold')
        self.figure.tight_layout()
        self.canvas.draw()

    def show_histogram(self):
        self.clear_axes()
        values = self.get_filtered_field(3) # Value column
        value_lengths = [len(v) for v in values]
       
        if value_lengths:
            n, bins, patches = self.ax.hist(value_lengths, bins=30, color='#12e0d6',
                                           edgecolor='#0f2f2f', linewidth=2, alpha=0.8)
           
            # Gradient colors for bars
            for i, patch in enumerate(patches):
                gradient_color = plt.cm.Blues(i / len(patches))
                patch.set_facecolor(gradient_color)
           
            self.ax.set_title('DISTRIBUTION DES LONGUEURS DES VALEURS XML', fontsize=18, pad=25, weight='bold')
            self.ax.set_xlabel('Longueur des Valeurs (caractères)', fontsize=14, weight='700')
            self.ax.set_ylabel('Fréquence', fontsize=14, weight='700')
            self.ax.grid(True, axis='y', linestyle='--', alpha=0.5, color='#FFFFFF')
           
            # Add statistics
            mean_length = np.mean(value_lengths)
            median_length = np.median(value_lengths)
           
            self.ax.axvline(mean_length, color='#f56565', linestyle='--', linewidth=3, alpha=0.9)
            self.ax.axvline(median_length, color='#12e0d6', linestyle='--', linewidth=3, alpha=0.9)
           
            self.ax.text(mean_length, max(n)*0.9, f'Moyenne: {mean_length:.1f}',
                        rotation=90, color='#f56565', fontsize=12, weight='bold')
            self.ax.text(median_length, max(n)*0.8, f'Médiane: {median_length:.1f}',
                        rotation=90, color='#12e0d6', fontsize=12, weight='bold')
        else:
            self.ax.text(0.5, 0.5, 'AUCUNE VALEUR DISPONIBLE',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=20, color='#f56565', weight='bold')
       
        self.figure.tight_layout()
        self.canvas.draw()

    def show_boxplot(self):
        self.clear_axes()
        values = self.get_filtered_field(3) # Value column
        value_lengths = [len(v) for v in values]
       
        if value_lengths:
            bp = self.ax.boxplot(value_lengths, vert=False, patch_artist=True,
                               boxprops=dict(facecolor='#12e0d6', color='#0f2f2f', alpha=0.8, linewidth=2),
                               whiskerprops=dict(color='#12e0d6', linewidth=3),
                               capprops=dict(color='#12e0d6', linewidth=3),
                               medianprops=dict(color='#FFFFFF', linewidth=4),
                               flierprops=dict(marker='o', markerfacecolor='#f56565', markersize=8))
           
            self.ax.set_title('ANALYSE STATISTIQUE DES LONGUEURS DES VALEURS XML', fontsize=18, pad=25, weight='bold')
            self.ax.set_xlabel('Longueur des Valeurs (caractères)', fontsize=14, weight='700')
            self.ax.grid(True, axis='x', linestyle='--', alpha=0.5, color='#FFFFFF')
           
            # Add comprehensive statistics
            median_val = np.median(value_lengths)
            q1_val = np.percentile(value_lengths, 25)
            q3_val = np.percentile(value_lengths, 75)
            min_val = np.min(value_lengths)
            max_val = np.max(value_lengths)
           
            stats_text = f'Min: {min_val}\nQ1: {q1_val:.1f}\nMédiane: {median_val:.1f}\nQ3: {q3_val:.1f}\nMax: {max_val}'
            self.ax.text(0.02, 0.98, stats_text, transform=self.ax.transAxes,
                        verticalalignment='top', bbox=dict(boxstyle='round,pad=0.5',
                        facecolor='#12e0d6', alpha=0.9, edgecolor='#0f2f2f'),
                        fontsize=12, color='#000000', weight='bold')
           
            mplcursors.cursor(bp['boxes']).connect("add",
                lambda sel: sel.annotation.set_text(f"Médiane: {median_val:.1f}\nQ1: {q1_val:.1f}\nQ3: {q3_val:.1f}\nIQR: {q3_val-q1_val:.1f}"))
        else:
            self.ax.text(0.5, 0.5, 'AUCUNE VALEUR DISPONIBLE',
                        ha='center', va='center', transform=self.ax.transAxes,
                        fontsize=20, color='#f56565', weight='bold')
       
        self.figure.tight_layout()
        self.canvas.draw()

    def export_chart(self):
        file_name, _ = QFileDialog.getSaveFileName(
            self, 'Exporter Graphique', '',
            'Images PNG (*.png);;Fichiers PDF (*.pdf);;Images SVG (*.svg);;Images JPEG (*.jpg)'
        )
        if file_name:
            try:
                self.figure.savefig(file_name, dpi=300, bbox_inches='tight',
                                  facecolor='#1a1a2e', edgecolor='none',
                                  pad_inches=0.2)
                print(f"[INFO] Chart exported to {file_name}")
                QMessageBox.information(self, '✅ SUCCÈS',
                                      f'Graphique exporté avec succès !\n\n📁 Emplacement:\n{file_name}')
            except Exception as e:
                print(f"[ERROR] Failed to export chart: {e}")
                QMessageBox.critical(self, '❌ ERREUR',
                                   f'Erreur lors de l\'exportation :\n\n{str(e)}')


class XmlParserApp(QWidget):
    def __init__(self, current_user=None):
        super().__init__()
        self.report_pdf_path = None
        self.current_user = current_user
        self.db_manager = OptionDatabaseManager()
        self.xml_parser = XMLParser(self.db_manager)
        self.xml_file_path = None
        self.xml_text_content = None
        self.csv_data = []
        self.qr_path = None
        self.initUI()

    def initUI(self):
        self.setWindowTitle('XML Parser Professional v2.0')
        self.setWindowIcon(QIcon("images/iconapp.ico"))
        self.setFixedSize(1000, 950)
        screen = QDesktopWidget().screenGeometry()
        self.setGeometry(
            (screen.width() - 1000) // 2,
            (screen.height() - 1000) // 2,
            1000, 1000
        )

        self.setStyleSheet("""
            QWidget {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #0f1419,
                    stop: 0.3 #1a1d29,
                    stop: 0.7 #252a3a,
                    stop: 1 #1a1f2e
                );
                font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
                color: #ffffff;
            }
            QFrame#headerFrame {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 rgba(18, 224, 214, 0.1),
                    stop: 0.5 rgba(15, 184, 176, 0.08),
                    stop: 1 rgba(12, 144, 136, 0.06)
                );
                border: 2px solid rgba(18, 224, 214, 0.3);
                border-radius: 12px;
                padding: 15px;
            }
            QLabel#titleLabel {
                font-size: 21px;
                font-weight: 700;
                color: #12e0d6;
                qproperty-alignment: 'AlignCenter';
                margin: 0;
                padding: 5px 0;
                text-shadow: 0 0 15px rgba(18, 224, 214, 0.5);
                letter-spacing: 1px;
                background: none;
                border: none;
                min-height: 30px;
            }
            QLabel#subtitleLabel {
                font-size: 13px;
                color: #a0b4c7;
                qproperty-alignment: 'AlignCenter';
                margin: 0;
                padding: 0 0 5px 0;
                background: none;
                border: none;
                min-height: 20px;
            }
            QLabel#sectionLabel {
                font-size: 14px;
                font-weight: 600;
                color: #12e0d6;
                margin: 8px 0 5px 0;
                padding: 0;
                background: none;
                border: none;
                text-shadow: 0 0 8px rgba(18, 224, 214, 0.3);
                min-height: 22px;
            }
            QLineEdit {
                background: rgba(18, 224, 214, 0.08);
                border: 1px solid rgba(18, 224, 214, 0.4);
                border-radius: 8px;
                padding: 8px 12px;
                color: #ffffff;
                font-size: 13px;
                min-height: 20px;
                max-height: 35px;
            }
            QLineEdit:focus {
                border: 1px solid #12e0d6;
                background: rgba(18, 224, 214, 0.12);
                box-shadow: 0 0 10px rgba(18, 224, 214, 0.3);
            }
            QLineEdit::placeholder {
                color: rgba(160, 180, 199, 0.7);
                font-style: italic;
            }
            QPushButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                border: none;
                border-radius: 8px;
                color: #000000;
                font-weight: 600;
                font-size: 11px;
                padding: 4px 8px;
                min-width: 160px;
                max-width: 160px;
                min-height: 40px;
                max-height: 40px;
            }
            QPushButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #14f2e8,
                    stop: 1 #12ccc4
                );
                box-shadow: 0 4px 15px rgba(18, 224, 214, 0.3);
                transform: translateY(-1px);
            }
            QPushButton:pressed {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #0d8f88,
                    stop: 1 #0b6f68
                );
                transform: translateY(0px);
            }
            QPushButton#executeButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #ff6b6b,
                    stop: 1 #e53e3e
                );
                color: #ffffff;
                font-size: 11px;
                font-weight: 700;
                padding: 4px 8px;
                min-width: 160px;
                max-width: 160px;
                min-height: 40px;
                max-height: 40px;
                border-radius: 8px;
                border: none;
            }
            QPushButton#executeButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #ff7675,
                    stop: 1 #fc5c65
                );
                box-shadow: 0 6px 20px rgba(255, 107, 107, 0.4);
            }
            QPushButton#dashboardButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #667eea,
                    stop: 1 #764ba2
                );
                color: #ffffff;
                font-size: 11px;
                font-weight: 700;
                padding: 4px 8px;
                min-width: 160px;
                max-width: 160px;
                min-height: 40px;
                max-height: 40px;
                border-radius: 8px;
                border: none;
            }
            QPushButton#dashboardButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #7c8ff0,
                    stop: 1 #8b5fb8
                );
                box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4);
            }
            QPushButton#searchButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #a855f7,
                    stop: 1 #8b5cf6
                );
                color: #ffffff;
                font-size: 11px;
                font-weight: 700;
                padding: 4px 8px;
                min-width: 160px;
                max-width: 160px;
                min-height: 40px;
                max-height: 40px;
                border-radius: 8px;
                border: none;
            }
            QPushButton#searchButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #c065f7,
                    stop: 1 #a36cf6
                );
                box-shadow: 0 6px 20px rgba(168, 85, 247, 0.4);
            }
            QPushButton#webButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #36d1dc,
                    stop: 1 #5b86e5
                );
                color: #ffffff;
                font-size: 11px;
                font-weight: 700;
                padding: 4px 8px;
                min-width: 160px;
                max-width: 160px;
                min-height: 40px;
                max-height: 40px;
                border-radius: 8px;
                border: none;
            }
            QPushButton#webButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #4ed9e4,
                    stop: 1 #6b96f5
                );
                box-shadow: 0 6px 20px rgba(54, 209, 220, 0.4);
            }
            QPushButton#exitButton {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #ef4444,
                    stop: 1 #dc2626
                );
                color: #ffffff;
                font-size: 11px;
                font-weight: 700;
                padding: 4px 8px;
                min-width: 160px;
                max-width: 160px;
                min-height: 40px;
                max-height: 40px;
                border-radius: 8px;
                border: none;
            }
            QPushButton#exitButton:hover {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #f87171,
                    stop: 1 #ef4444
                );
                box-shadow: 0 6px 20px rgba(239, 68, 68, 0.4);
            }
            QRadioButton {
                color: #a0b4c7;
                font-size: 13px;
                font-weight: 500;
                spacing: 8px;
                padding: 8px 0;
                background: transparent;
                border: none;
                margin: 0;
                min-height: 24px;
            }
            QRadioButton::indicator {
                width: 16px;
                height: 16px;
                border-radius: 8px;
                border: 1px solid #12e0d6;
                background: rgba(18, 224, 214, 0.05);
            }
            QRadioButton::indicator:checked {
                background: qradialgradient(
                    cx: 0.5, cy: 0.5, radius: 0.5,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                border: 1px solid #0fb8b0;
            }
            QRadioButton:hover {
                color: #ffffff;
            }
            QRadioButton::text {
                color: #a0b4c7;
                padding-left: 5px;
            }
            QRadioButton:hover::text {
                color: #ffffff;
            }
            QTextEdit {
                background: rgba(18, 224, 214, 0.06);
                border: 1px solid rgba(18, 224, 214, 0.3);
                border-radius: 8px;
                padding: 10px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 12px;
                color: #ffffff;
                selection-background-color: #12e0d6;
                selection-color: #000000;
                min-height: 100px;
                max-height: 200px;
            }
            QTextEdit:focus {
                border: 1px solid #12e0d6;
                background: rgba(18, 224, 214, 0.10);
            }
            QFrame#cardFrame {
                background: rgba(18, 224, 214, 0.06);
                border: 1px solid rgba(18, 224, 214, 0.2);
                border-radius: 20px;
                padding: 15px;
                margin: 5px;
            }
            QGroupBox {
                font-size: 13px;
                font-weight: 600;
                color: #12e0d6;
                border: 1px solid rgba(18, 224, 214, 0.3);
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background: rgba(18, 224, 214, 0.04);
            }
            QGroupBox::title {
                color: #12e0d6;
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0px;
                background: transparent;
                border: none;
                margin-left: 8px;
            }
            QProgressBar {
                border: 2px solid #12e0d6;
                border-radius: 8px;
                background: rgba(18, 224, 214, 0.1);
                text-align: center;
                color: #ffffff;
                font-size: 12px;
                font-weight: 600;
                min-height: 20px;
                max-height: 25px;
            }
            QProgressBar::chunk {
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #12e0d6,
                    stop: 1 #0fb8b0
                );
                border-radius: 6px;
            }
            QScrollBar:vertical {
                background: rgba(255, 255, 255, 0.05);
                width: 12px;
                border-radius: 6px;
                margin: 0;
            }
            QScrollBar::handle:vertical {
                background: rgba(18, 224, 214, 0.6);
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(18, 224, 214, 0.8);
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0;
            }
            QComboBox {
                background: rgba(18, 224, 214, 0.08);
                border: 1px solid rgba(18, 224, 214, 0.4);
                border-radius: 6px;
                padding: 6px 10px;
                color: #ffffff;
                font-size: 12px;
                min-height: 20px;
                max-height: 30px;
            }
            QComboBox:hover {
                border: 1px solid #12e0d6;
                background: rgba(18, 224, 214, 0.12);
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
                background: transparent;
            }
            QComboBox::down-arrow {
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #12e0d6;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                background: #1a1f2e;
                border: 1px solid #12e0d6;
                border-radius: 6px;
                color: #ffffff;
                selection-background-color: rgba(18, 224, 214, 0.3);
                padding: 4px;
            }
            QLabel#logo_label {
                background: transparent;
                border: none;
                padding: 0;
                margin: 0;
            }
        """)

        self.main_layout = QVBoxLayout()
        self.main_layout.setSpacing(15)
        self.main_layout.setContentsMargins(20, 15, 20, 15)
        self.setLayout(self.main_layout)

        # Header
        header_frame = QFrame()
        header_frame.setObjectName('headerFrame')
        header_layout = QVBoxLayout(header_frame)
        header_layout.setSpacing(12)
        header_layout.setContentsMargins(10, 10, 10, 10)

        self.logo_label = QLabel(self)
        self.logo_label.setObjectName('logo_label')
        pixmap = QPixmap('images/logo.png').scaled(200, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        if pixmap.isNull():
            self.logo_label.setText("🚀 XML PARSER PRO")
            self.logo_label.setObjectName('titleLabel')
        else:
            self.logo_label.setPixmap(pixmap)
        self.logo_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(self.logo_label)

        title_label = QLabel('XML Parser Professional')
        title_label.setObjectName('titleLabel')
        title_label.setAlignment(Qt.AlignCenter)

        subtitle_label = QLabel('✨ Analysez et convertissez vos fichiers XML en rapports professionnels de haute qualité ✨')
        subtitle_label.setObjectName('subtitleLabel')
        subtitle_label.setAlignment(Qt.AlignCenter)

        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)

        separator = QFrame()
        separator.setObjectName('separatorLine')
        separator.setFrameShape(QFrame.HLine)
        separator.setFixedHeight(3)

        # File selection section
        file_card = QFrame()
        file_card.setObjectName('cardFrame')
        file_layout = QVBoxLayout(file_card)
        file_layout.setSpacing(20)

        file_section_label = QLabel('📁 Sélection du Fichier XML')
        file_section_label.setObjectName('sectionLabel')

        self.filePathInput = QLineEdit()
        self.filePathInput.setPlaceholderText('✨ Sélectionnez votre fichier XML pour commencer l\'analyse...')
        self.filePathInput.setFixedHeight(65)

        self.btnSelectFile = QPushButton('📂 Parcourir les Fichiers')
        self.btnSelectFile.setFixedSize(220, 65)
        self.btnSelectFile.clicked.connect(self.showDialog)

        file_input_layout = QHBoxLayout()
        file_input_layout.addWidget(self.filePathInput, 1)
        file_input_layout.addSpacing(20)
        file_input_layout.addWidget(self.btnSelectFile)

        file_layout.addWidget(file_section_label)
        file_layout.addLayout(file_input_layout)

        # Report type section
        report_card = QFrame()
        report_card.setObjectName('cardFrame')
        report_layout = QVBoxLayout(report_card)
        report_layout.setSpacing(20)

        report_section_label = QLabel('📊 Configuration du Type de Rapport')
        report_section_label.setObjectName('sectionLabel')

        self.report_group = QButtonGroup()
        self.typrep1 = QRadioButton('📋 Rapport Standard - Analyse rapide et concise')
        self.typrep1.setChecked(True)
        self.typrep2 = QRadioButton('📈 Rapport Détaillé - Analyse complète et approfondie')
        self.report_group.addButton(self.typrep1, 1)
        self.report_group.addButton(self.typrep2, 2)

        radio_layout = QVBoxLayout()
        radio_layout.setSpacing(15)
        radio_layout.addWidget(self.typrep1)
        radio_layout.addWidget(self.typrep2)

        report_layout.addWidget(report_section_label)
        report_layout.addLayout(radio_layout)

        # Action buttons
        BUTTON_WIDTH = 160
        BUTTON_HEIGHT = 40

        main_buttons_layout = QVBoxLayout()
        main_buttons_layout.setSpacing(20)
        main_buttons_layout.setAlignment(Qt.AlignCenter)

        # Primary buttons row (Générer Rapport and Tableau de Bord)
        primary_buttons_layout = QHBoxLayout()
        primary_buttons_layout.setSpacing(20)
        primary_buttons_layout.setAlignment(Qt.AlignCenter)

        self.btnRunScript = QPushButton('🚀 GÉNÉRER LE RAPPORT')
        self.btnRunScript.setObjectName('executeButton')
        self.btnRunScript.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
        self.btnRunScript.clicked.connect(self.runScript)
        primary_buttons_layout.addWidget(self.btnRunScript)

        self.btnDashboard = QPushButton('📊 TABLEAU DE BORD')
        self.btnDashboard.setObjectName('dashboardButton')
        self.btnDashboard.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
        self.btnDashboard.clicked.connect(self.show_dashboard)
        primary_buttons_layout.addWidget(self.btnDashboard)

        # Secondary buttons row (Recherche Avancée and Interface Web)
        secondary_buttons_layout = QHBoxLayout()
        secondary_buttons_layout.setSpacing(20)
        secondary_buttons_layout.setAlignment(Qt.AlignCenter)

        self.search_button = QPushButton('🔍 RECHERCHE AVANCÉE')
        self.search_button.setObjectName('searchButton')
        self.search_button.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
        self.search_button.clicked.connect(self.open_search_dialog)
        secondary_buttons_layout.addWidget(self.search_button)

        # Exit button row
        exit_buttons_layout = QHBoxLayout()
        exit_buttons_layout.setAlignment(Qt.AlignCenter)

        self.back_button = QPushButton('⬅️ QUITTER L\'APPLICATION')
        self.back_button.setObjectName('exitButton')
        self.back_button.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
        self.back_button.clicked.connect(self.back_to_login)
        exit_buttons_layout.addWidget(self.back_button)

        # Assemble button layouts
        main_buttons_layout.addLayout(primary_buttons_layout)
        main_buttons_layout.addLayout(secondary_buttons_layout)
        main_buttons_layout.addLayout(exit_buttons_layout)

        # Logs section
        logs_group = QGroupBox("📝 Journal de Traitement en Temps Réel")
        logs_layout = QVBoxLayout()
        self.logs_text = QTextEdit()
        self.logs_text.setMaximumHeight(220)
        self.logs_text.setReadOnly(True)
        self.logs_text.setPlaceholderText("Les logs de traitement apparaîtront ici en temps réel...")
        logs_layout.addWidget(self.logs_text)
        logs_group.setLayout(logs_layout)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedSize(500, 45)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("⚡ Prêt pour le traitement... %p%")
        self.progress_bar.setVisible(False)

        # Assemble layout
        self.main_layout.addWidget(header_frame)
        self.main_layout.addWidget(separator)
        self.main_layout.addSpacing(20)
        self.main_layout.addWidget(file_card)
        self.main_layout.addSpacing(20)
        self.main_layout.addWidget(report_card)
        self.main_layout.addSpacing(30)
        self.main_layout.addWidget(logs_group)
        self.main_layout.addSpacing(15)
        self.main_layout.addWidget(self.progress_bar, alignment=Qt.AlignCenter)
        self.main_layout.addSpacing(20)
        self.main_layout.addLayout(main_buttons_layout)
        self.main_layout.addStretch()

        self.setLayout(self.main_layout)

    def back_to_login(self):
        self.close()
        self.login_window = LoginWindow()
        self.login_window.show()

    def open_search_dialog(self):
        try:
            self.log_message("Opening search dialog")
            dialog = OptionSearchDialog(self.db_manager, self)
            dialog.exec_()
            self.log_message("Search dialog closed")
        except Exception as e:
            self.log_message(f"Error opening search dialog: {e}")
            QMessageBox.critical(self, "Error", f"Error opening search dialog: {e}")

    def showDialog(self):
        fname, _ = QFileDialog.getOpenFileName(self, 'Sélectionner le fichier XML', '', 'Fichiers XML (*.xml);;Tous les fichiers (*)')
        if fname:
            self.log_message(f"Selected file: {fname}")
            self.xml_file_path = fname
            self.filePathInput.setText(fname)

    def log_message(self, message):
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.logs_text.append(f"[{timestamp}] {message}")
            QApplication.processEvents()
        except Exception as e:
            print(f"Error logging message: {e}")

    def show_dashboard(self):
        if not self.csv_data:
            self.log_message("Warning: No XML file processed. Generate a report first.")
            QMessageBox.warning(self, 'Attention', 'Aucun fichier XML traité. Veuillez générer un rapport d\'abord.')
            return
        self.log_message("Opening dashboard")
        dashboard_window = DashboardWindow(self.csv_data, self.qr_path, self)
        dashboard_window.exec_()
        self.log_message("Dashboard closed")

    def runScript(self):
        if not self.xml_file_path:
            self.log_message("Warning: No XML file selected")
            QMessageBox.warning(self, 'Attention', 'Veuillez d\'abord sélectionner un fichier XML.')
            return

        self.log_message(f"Starting script execution for file: {self.xml_file_path}")
        fname = self.xml_file_path
        type_etat = 2 if self.typrep2.isChecked() else 1
        self.csv_data = []
        xml_text_lines = []
        option_counts = Counter()

        # Initialize progress bar
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Analyse du fichier XML... %p%")
        QApplication.processEvents()

        try:
            # Parse XML file directly (UNCHANGED PARSING LOGIC)
            tree = ET.parse(fname)
            root = tree.getroot()

            variable_list = []
            nb_option = 0
            parsed_data = {'options': [], 'variables': [], 'filepath': fname}
            current_option = None

            for element in root.iter():
                if element.tag == 'option':
                    option_name = element.get('name')
                    procedure = element.get('prog_name') or ''
                    if option_name and option_name not in {'MAJPROG', 'MAJMENU', 'FORPCT', 'MAJOPT'}:
                        nb_option += 1
                        option_counts[option_name] += 1
                        current_option = {
                            'name': option_name,
                            'procedure': procedure,
                            'variables': [],
                            'is_covered': True,
                            'coverage_details': 'Direct parsing'
                        }
                        parsed_data['options'].append(current_option)
                        xml_text_lines.append(f"Option: {option_name}")
                        xml_text_lines.append(f"  Procédure: {procedure}")
                        if procedure:
                            if len(procedure) > 30:
                                procedure = procedure[:15] + '\n' + procedure[15:30] + '\n' + procedure[30:]
                            elif len(procedure) > 15:
                                procedure = procedure[:15] + '\n' + procedure[15:]

                elif element.tag in ['set_variable', 'field_input']:
                    variable_name = element.get('name')
                    value = element.get('value', '')
                    if variable_name and value and value not in {'@SKIP@', '@DOWN@', '@QUIT@'} and variable_name not in {'code_option', 'nom_procedure', 'nature_option_saisie'} and value[:7] != 'NEWPORT':
                        if element.tag == 'field_input' and variable_name[:14] != 'TABLEAU_OPTION':
                            variable = {'name': variable_name, 'value': value, 'type': element.tag, 'is_required': False}
                            if current_option:
                                current_option['variables'].append(variable)
                            parsed_data['variables'].append(variable)
                            xml_text_lines.append(f"  Variable: {variable_name} = {value}")
                            variable_list.append((variable_name, value))

                elif element.tag == 'from_variable':
                    for variable in variable_list:
                        if element.get('name') == variable[0] and variable[1] not in {'@SKIP@', '@DOWN@', '@QUIT@'}:
                            if current_option:
                                current_option['variables'].append({'name': variable[0], 'value': variable[1], 'type': 'from_variable', 'is_required': False})
                            parsed_data['variables'].append({'name': variable[0], 'value': variable[1], 'type': 'from_variable', 'is_required': False})
                            xml_text_lines.append(f"  Variable: {variable[0]} = {variable[1]}")
                            break
                    variable_list.clear()

            # Log parsed data for debugging
            self.log_message(f"Parsed data summary: {len(parsed_data['options'])} options, {len(parsed_data['variables'])} variables")
            self.log_message(f"Sample options: {parsed_data['options'][:3]}")
            self.log_message(f"Sample variables: {parsed_data['variables'][:3]}")

            total_elements = len(parsed_data['options']) + len(parsed_data['variables'])
            processed_elements = len(parsed_data['options'])
            self.progress_bar.setValue(5)

            self.progress_bar.setFormat("Sauvegarde dans la base de données... %p%")
            self.progress_bar.setValue(10)
            QApplication.processEvents()

            if not self.save_to_database(parsed_data, fname):
                self.log_message("Error: Failed to save data to database")
                self.progress_bar.setVisible(False)
                QMessageBox.critical(self, 'Erreur', 'Échec de la sauvegarde dans la base de données.')
                return

            self.progress_bar.setValue(50)

            # Generate file names
            base_name = os.path.basename(self.xml_file_path).replace('.xml', '').replace(' ', '_')
            report_csv_name = os.path.join("reports", f"{base_name}.csv")
            report_excel_name = os.path.join("reports", f"{base_name}.xlsx")
            report_pdf_name = os.path.join("reports", f"{base_name}.pdf")


            # Build csv_data
            for option in parsed_data['options']:
                procedure = option['procedure']
                option_name = option['name']
                for variable in option.get('variables', []):
                    var_name = variable['name']
                    var_value = variable['value']
                    # Include all parsed variables, no redundant filtering
                    self.csv_data.append([option_name, procedure, var_name, var_value])
                    self.log_message(f"CSV row added: {option_name}, {procedure}, {var_name}, {var_value}")
                    option_name = ''
                    procedure = ''
                if not option.get('variables', []):
                    self.csv_data.append([option_name, procedure, '', ''])
                    self.log_message(f"CSV row added: {option_name}, {procedure}, '', ''")

            # Generate CSV
            self.progress_bar.setFormat("Génération du CSV... %p%")
            self.progress_bar.setValue(85)
            QApplication.processEvents()

            os.makedirs("reports", exist_ok=True)
            with open(report_csv_name, 'w', newline='', encoding='utf-8') as csv_file:
                csv_writer = csv.writer(csv_file, delimiter=';')
                csv_writer.writerow(['Option', 'Procédure', 'Variable', 'Valeur'])
                for row in self.csv_data:
                    csv_writer.writerow(row)
            self.log_message(f"Generated CSV: {report_csv_name}")

            # Generate Excel
            self.progress_bar.setFormat("Génération du Excel... %p%")
            self.progress_bar.setValue(90)
            QApplication.processEvents()
            report_generator = ReportGenerator(parsed_data, base_name, log_func=self.log_message)
            excel_path = report_generator.generate_excel(report_excel_name)
            if not excel_path:
                self.log_message("Error: Failed to generate Excel")
                self.progress_bar.setVisible(False)
                QMessageBox.critical(self, 'Erreur', 'Échec de la génération du fichier Excel.')
                return
            self.log_message(f"Excel file generated: {excel_path}")

            # Generate PDF
            self.progress_bar.setFormat("Génération du PDF... %p%")
            self.progress_bar.setValue(95)
            QApplication.processEvents()

            report_generator = ReportGenerator(parsed_data, base_name, log_func=self.log_message)
            self.report_pdf_path = report_generator.generate_pdf(type_etat=type_etat)
            if not self.report_pdf_path:
                self.log_message("Error: Failed to generate PDF")
                self.progress_bar.setVisible(False)
                QMessageBox.critical(self, 'Erreur', 'Échec de la génération du fichier PDF.')
                return

            # Generate QR code
            self.progress_bar.setFormat("Génération du QR code... %p%")
            self.progress_bar.setValue(98)
            QApplication.processEvents()

            self.xml_text_content = "\n".join([f"{option}: {count}" for option, count in option_counts.items()])
            if len(self.xml_text_content) > 2000:
                self.log_message("Warning: XML content is long. Truncating to 2000 characters.")
                self.xml_text_content = self.xml_text_content[:2000] + "\n[Texte tronqué...]"
            self.generate_qr_code()

            self.progress_bar.setFormat("Terminé ! %p%")
            self.progress_bar.setValue(100)
            QApplication.processEvents()

            self.log_message(f"Processing complete: {nb_option} options found, PDF: {self.report_pdf_path}")
            msg = QMessageBox(self)
            msg.setWindowTitle('✅ Traitement Terminé')
            msg.setText(f'Le traitement est terminé avec succès !\n\n📊 {nb_option} options trouvées\n📄 Rapport PDF: {self.report_pdf_path}\n📱 QR code: Contient le texte du XML')
            msg.setIcon(QMessageBox.Information)
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #1e293b;
                    color: #e2e8f0;
                }
                QMessageBox QPushButton {
                    background-color: #22d3ee;
                    border: none;
                    border-radius: 8px;
                    padding: 8px 16px;
                    color: #1e293b;
                    font-weight: bold;
                }
                QMessageBox QPushButton:hover {
                    background-color: #06b6d4;
                }
            """)
            msg.exec_()

            self.progress_bar.setVisible(False)

            if os.path.exists(self.report_pdf_path):
                self.log_message("PDF file generated successfully")
                self.showPDFReport()
            else:
                self.log_message("Error: PDF file not generated")
                QMessageBox.critical(self, 'Erreur', f'Le fichier PDF {self.report_pdf_path} n\'a pas été généré.')

        except Exception as e:
            self.log_message(f"Error during XML processing: {e}")
            self.progress_bar.setVisible(False)
            QMessageBox.critical(self, 'Erreur', f'Erreur lors de l\'analyse du fichier XML: {e}')

    def showPDFReport(self):
        if self.report_pdf_path:
            try:
                Popen([self.report_pdf_path], shell=True)
            except Exception as e:
                self.log_message(f"Erreur lors de l'ouverture du PDF : {e}")
        else:
            self.log_message('Aucun rapport PDF généré.')
            QMessageBox.critical(self, 'Erreur', 'Aucun rapport PDF généré.')

    def save_to_database(self, parsed_data, filepath):
        """Save parsed XML data to the database."""
        try:
            self.log_message(f"Starting database save for file: {filepath}")
            # Calculate file hash
            with open(filepath, 'rb') as f:
                file_content = f.read()
                file_hash = hashlib.sha256(file_content).hexdigest()
            self.log_message(f"Calculated file hash: {file_hash}")

            # Add XML file to database
            xml_file_id = self.db_manager.add_xml_file(os.path.basename(filepath), filepath, file_hash)
            self.log_message(f"Added XML file with ID: {xml_file_id}")
            if not xml_file_id:
                self.log_message("Error: Failed to add XML file to database")
                return False

            # Add options, variables, and treatment codes
            option_count = 0
            for option in parsed_data['options']:
                option_name = option['name']
                procedure = option['procedure']
                is_covered = option.get('is_covered', False)
                coverage_details = option.get('coverage_details', '')

                self.log_message(f"Adding option: {option_name}, procedure: {procedure}")
                option_id = self.db_manager.add_option(
                    xml_file_id, option_name, procedure, is_covered, procedure, coverage_details
                )
                if not option_id:
                    self.log_message(f"Warning: Option {option_name} not added (possible duplicate)")
                    continue

                option_count += 1

            self.log_message(f"Database save completed: {option_count} options")
            return True
        except sqlite3.Error as e:
            self.log_message(f"Database save error: {str(e)}")
            return False
        except Exception as e:
            self.log_message(f"Unexpected error during database save: {str(e)}")
            return False

    def generate_qr_code(self):
        try:
            if not self.xml_text_content:
                self.log_message("Error: No XML content to encode in QR code")
                QMessageBox.critical(self, 'Erreur', 'Aucun contenu XML à encoder dans le QR code.')
                return

            self.log_message(f"Generating QR code for content (length: {len(self.xml_text_content)})")
            qr = qrcode.QRCode(
                version=10,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4
            )
            qr.add_data(self.xml_text_content)
            qr.make(fit=True)

            img = qr.make_image(fill='black', back_color='white')
            self.qr_path = os.path.join("reports", 'xml_option_qrcode.png')
            os.makedirs("reports", exist_ok=True)
            img.save(self.qr_path)
            self.log_message(f"QR code saved: {self.qr_path}")

        except qrcode.exceptions.DataOverflowError:
            self.log_message("Error: XML content too long for QR code")
            QMessageBox.critical(self, 'Erreur', 'Le contenu XML est trop long pour être encodé dans un QR code.')
        except Exception as e:
            self.log_message(f"Error generating QR code: {e}")
            QMessageBox.critical(self, 'Erreur', f'Erreur lors de la génération du QR code : {str(e)}')

    def show_qr_code(self, image_path):
        qr_window = QWidget()
        qr_window.setWindowTitle("QR Code du contenu XML")
        qr_window.setFixedSize(400, 400)
        qr_window.setStyleSheet("""
            QWidget {
                background: #1e293b;
            }
            QLabel {
                color: #e2e8f0;
            }
        """)

        layout = QVBoxLayout()
        label = QLabel()
        pixmap = QPixmap(image_path).scaled(350, 350, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        if pixmap.isNull():
            label.setText("Erreur : Impossible de charger l'image QR.")
            label.setStyleSheet("color: #ef4444;")
        else:
            label.setPixmap(pixmap)
        label.setAlignment(Qt.AlignCenter)

        layout.addWidget(label)
        qr_window.setLayout(layout)
        qr_window.show()

        self.qr_window = qr_window



import smtplib
import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage


import smtplib
import os
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from PyQt5.QtCore import QThread, pyqtSignal

class EmailManager:
    """Gère l'envoi d'emails via SMTP avec design professionnel."""
    
    def __init__(self):
        """Initialise le gestionnaire d'emails."""
        self.smtp_server = None
        self.smtp_port = None
        self.sender_email = None
        self.sender_password = None
        self.app_name = "Application"
        print("[DEBUG] EmailManager initialized")
    
    def configure_smtp(self, smtp_server, smtp_port, sender_email, sender_password, app_name):
        """Configure les paramètres SMTP."""
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.sender_email = sender_email
        self.sender_password = sender_password
        self.app_name = app_name
        print(f"[DEBUG] EmailManager: Configured SMTP with {smtp_server}:{smtp_port}, sender={sender_email}")
    
    def _create_email_template(self, username, email_type="confirmation"):
        """Crée le template HTML professionnel pour l'email."""
        
        if email_type == "confirmation":
            title = f"Bienvenue sur {self.app_name}!"
            header_title = "Votre Application"
            welcome_title = f"Bienvenue, {username}!"
            welcome_subtitle = "Nous sommes ravis de vous compter parmi nous"
            card_title = "🎉 Inscription réussie"
            card_content = f"""
                <p>Votre compte a été créé avec succès sur notre plateforme.</p>
                <p>Nous avons reçu votre demande et notre équipe procède actuellement à la validation de votre compte.</p>
                <div class="status-badge">⏳ En attente de validation</div>
            """
            cta_content = '<p style="color: #6b7280; margin-bottom: 20px;">Vous recevrez un email de confirmation dès que votre compte sera activé.</p>'
            footer_title = "Merci de votre confiance"
            footer_content = "Notre équipe vous souhaite la bienvenue et espère que vous apprécierez votre expérience avec nous."
            
        elif email_type == "acceptance":
            title = f"Votre compte a été validé sur {self.app_name}"
            header_title = "Compte Activé!"
            welcome_title = f"Félicitations, {username}!"
            welcome_subtitle = "Votre compte a été validé avec succès"
            card_title = "✅ Compte activé"
            card_content = f"""
                <p>Votre compte a été validé par notre équipe d'administration.</p>
                <p>Vous pouvez maintenant vous connecter et profiter pleinement de tous nos services.</p>
                <div class="status-badge accepted-badge">✓ Compte actif</div>
            """
            cta_content = '<a href="#" class="cta-button">Se connecter maintenant</a>'
            footer_title = "Bienvenue dans notre communauté!"
            footer_content = "Nous sommes impatients de vous voir utiliser notre plateforme."
        
        current_year = datetime.now().year
        
        html_content = f"""
        <!DOCTYPE html>
        <html lang="fr">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{title}</title>
            <style>
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background-color: #f8f9fa;
                    padding: 20px;
                }}
                
                .email-container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background-color: #ffffff;
                    border-radius: 12px;
                    overflow: hidden;
                    box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
                }}
                
                .header {{
                    background: linear-gradient(135deg, #0ed1b2 0%, #0ab896 100%);
                    padding: 40px 30px;
                    text-align: center;
                    position: relative;
                }}
                
                .header::before {{
                    content: '';
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    bottom: 0;
                    background: url('data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"><defs><pattern id="grain" patternUnits="userSpaceOnUse" width="100" height="100"><circle cx="20" cy="20" r="1" fill="white" opacity="0.1"/><circle cx="80" cy="40" r="1" fill="white" opacity="0.1"/><circle cx="40" cy="70" r="1" fill="white" opacity="0.1"/></pattern></defs><rect width="100" height="100" fill="url(%23grain)"/></svg>');
                }}
                
                .logo-container {{
                    position: relative;
                    z-index: 2;
                    margin-bottom: 20px;
                }}
                
                .logo {{
                    width: 120px;
                    height: 120px;
                    background-color: #ffffff;
                    border-radius: 20px;
                    display: inline-flex;
                    align-items: center;
                    justify-content: center;
                    box-shadow: 0 6px 25px rgba(0, 0, 0, 0.2);
                    padding: 15px;
                }}
                
                .logo img {{
                    width: 100%;
                    height: 100%;
                    object-fit: contain;
                }}
                
                .header-title {{
                    color: #ffffff;
                    font-size: 32px;
                    font-weight: 700;
                    margin: 15px 0 0 0;
                    position: relative;
                    z-index: 2;
                    text-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
                }}
                
                .content {{
                    padding: 40px 30px;
                    background-color: #ffffff;
                }}
                
                .welcome-section {{
                    text-align: center;
                    margin-bottom: 40px;
                }}
                
                .welcome-title {{
                    color: #22211f;
                    font-size: 28px;
                    font-weight: 600;
                    margin-bottom: 15px;
                }}
                
                .welcome-subtitle {{
                    color: #6b7280;
                    font-size: 18px;
                    line-height: 1.6;
                }}
                
                .info-card {{
                    background: linear-gradient(135deg, #f8fdfc 0%, #ffffff 100%);
                    border: 2px solid #0ed1b2;
                    border-radius: 12px;
                    padding: 30px;
                    margin: 30px 0;
                    position: relative;
                    overflow: hidden;
                }}
                
                .info-card::before {{
                    content: '';
                    position: absolute;
                    top: 0;
                    left: 0;
                    width: 4px;
                    height: 100%;
                    background: linear-gradient(180deg, #0ed1b2 0%, #0ab896 100%);
                }}
                
                .info-card h3 {{
                    color: #22211f;
                    font-size: 20px;
                    font-weight: 600;
                    margin-bottom: 15px;
                }}
                
                .info-card p {{
                    color: #4b5563;
                    font-size: 16px;
                    line-height: 1.6;
                    margin-bottom: 10px;
                }}
                
                .status-badge {{
                    display: inline-block;
                    background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
                    color: #92400e;
                    padding: 10px 20px;
                    border-radius: 25px;
                    font-size: 14px;
                    font-weight: 600;
                    margin-top: 15px;
                    border: 1px solid #f59e0b;
                }}
                
                .cta-section {{
                    text-align: center;
                    margin: 40px 0;
                }}
                
                .cta-button {{
                    display: inline-block;
                    background: linear-gradient(135deg, #0ed1b2 0%, #0ab896 100%);
                    color: #ffffff;
                    text-decoration: none;
                    padding: 16px 35px;
                    border-radius: 10px;
                    font-weight: 600;
                    font-size: 18px;
                    box-shadow: 0 4px 15px rgba(14, 209, 178, 0.3);
                    transition: all 0.3s ease;
                }}
                
                .cta-button:hover {{
                    transform: translateY(-2px);
                    box-shadow: 0 6px 25px rgba(14, 209, 178, 0.4);
                }}
                
                .footer {{
                    background-color: #22211f;
                    color: #ffffff;
                    padding: 30px;
                    text-align: center;
                }}
                
                .footer-content {{
                    margin-bottom: 20px;
                }}
                
                .footer h4 {{
                    font-size: 20px;
                    font-weight: 600;
                    margin-bottom: 15px;
                    color: #0ed1b2;
                }}
                
                .footer p {{
                    font-size: 16px;
                    line-height: 1.6;
                    color: #d1d5db;
                }}
                
                .social-links {{
                    margin: 20px 0;
                }}
                
                .social-link {{
                    display: inline-block;
                    width: 45px;
                    height: 45px;
                    background-color: #0ed1b2;
                    border-radius: 50%;
                    margin: 0 10px;
                    text-decoration: none;
                    color: #ffffff;
                    line-height: 45px;
                    text-align: center;
                    font-size: 18px;
                    transition: all 0.3s ease;
                }}
                
                .social-link:hover {{
                    background-color: #0ab896;
                    transform: translateY(-2px);
                }}
                
                .copyright {{
                    font-size: 14px;
                    color: #9ca3af;
                    border-top: 1px solid #374151;
                    padding-top: 20px;
                    margin-top: 20px;
                }}
                
                .divider {{
                    width: 60px;
                    height: 4px;
                    background: linear-gradient(135deg, #0ed1b2 0%, #0ab896 100%);
                    margin: 25px auto;
                    border-radius: 2px;
                }}

                /* Version pour email d'acceptation */
                .accepted-badge {{
                    background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
                    color: #065f46;
                    border: 1px solid #10b981;
                }}
                
                @media (max-width: 600px) {{
                    body {{
                        padding: 10px;
                    }}
                    
                    .header {{
                        padding: 30px 20px;
                    }}
                    
                    .logo {{
                        width: 100px;
                        height: 100px;
                    }}
                    
                    .header-title {{
                        font-size: 26px;
                    }}
                    
                    .content {{
                        padding: 30px 20px;
                    }}
                    
                    .footer {{
                        padding: 25px 20px;
                    }}
                    
                    .welcome-title {{
                        font-size: 24px;
                    }}
                    
                    .cta-button {{
                        padding: 14px 25px;
                        font-size: 16px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <!-- Header -->
                <div class="header">
                    <div class="logo-container">
                        <div class="logo">
                            <img src="cid:neoxam" alt="Logo {self.app_name}" />
                        </div>
                    </div>
                    <h1 class="header-title">{header_title}</h1>
                </div>
                
                <!-- Content -->
                <div class="content">
                    <div class="welcome-section">
                        <h2 class="welcome-title">{welcome_title}</h2>
                        <p class="welcome-subtitle">{welcome_subtitle}</p>
                        <div class="divider"></div>
                    </div>
                    
                    <div class="info-card">
                        <h3>{card_title}</h3>
                        {card_content}
                    </div>
                    
                    <div class="cta-section">
                        {cta_content}
                    </div>
                </div>
                
                <!-- Footer -->
                <div class="footer">
                    <div class="footer-content">
                        <h4>{footer_title}</h4>
                        <p>{footer_content}</p>
                    </div>
                    
                    <div class="social-links">
                        <a href="#" class="social-link">📧</a>
                        <a href="#" class="social-link">🌐</a>
                        <a href="#" class="social-link">📱</a>
                    </div>
                    
                    <div class="copyright">
                        <p>&copy; {current_year} {self.app_name}. Tous droits réservés.</p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html_content
    
    def send_confirmation_email(self, recipient_email, username, logo_path=None):
        """Envoie un email de confirmation d'inscription avec design professionnel."""
        print(f"[DEBUG] EmailManager: Sending confirmation email to {recipient_email}")
        try:
            # Handle logo path for PyInstaller
            if logo_path:
                if hasattr(sys, '_MEIPASS'):
                    logo_path = os.path.join(sys._MEIPASS, logo_path)
                if not os.path.exists(logo_path):
                    print(f"[ERROR] EmailManager: Logo file not found: {logo_path}")
                    logo_path = None

            msg = MIMEMultipart('related')
            msg['From'] = f"{self.app_name} <{self.sender_email}>"
            msg['To'] = recipient_email
            msg['Subject'] = f"Bienvenue sur {self.app_name}!"
            
            html_content = self._create_email_template(username, "confirmation")
            msg.attach(MIMEText(html_content, 'html'))
            
            if logo_path and os.path.exists(logo_path):
                with open(logo_path, 'rb') as img_file:
                    img = MIMEImage(img_file.read())
                    img.add_header('Content-ID', '<neoxam>')
                    img.add_header('Content-Disposition', 'inline', filename=os.path.basename(logo_path))
                    msg.attach(img)
                    print(f"[DEBUG] EmailManager: Attached logo {logo_path} to email")
            
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)
            
            print(f"[INFO] EmailManager: Confirmation email sent to {recipient_email}")
            return True, "Email de confirmation envoyé."
        except smtplib.SMTPAuthenticationError as e:
            print(f"[ERROR] EmailManager: SMTP authentication error: {e}")
            return False, "Erreur d'authentification SMTP. Vérifiez l'email et le mot de passe."
        except smtplib.SMTPException as e:
            print(f"[ERROR] EmailManager: SMTP error sending confirmation email: {e}")
            return False, f"Erreur SMTP: {str(e)}"
        except Exception as e:
            print(f"[ERROR] EmailManager: General error sending confirmation email: {e}")
            return False, f"Erreur lors de l'envoi de l'email: {str(e)}"
    
    def send_acceptance_email(self, recipient_email, username, logo_path=None):
        """Envoie un email de validation de compte avec design professionnel."""
        print(f"[DEBUG] EmailManager: Sending acceptance email to {recipient_email}")
        try:
            # Handle logo path for PyInstaller
            if logo_path:
                if hasattr(sys, '_MEIPASS'):
                    logo_path = os.path.join(sys._MEIPASS, logo_path)
                if not os.path.exists(logo_path):
                    print(f"[ERROR] EmailManager: Logo file not found: {logo_path}")
                    logo_path = None

            msg = MIMEMultipart('related')
            msg['From'] = f"{self.app_name} <{self.sender_email}>"
            msg['To'] = recipient_email
            msg['Subject'] = f"Votre compte a été validé sur {self.app_name}"
            
            html_content = self._create_email_template(username, "acceptance")
            msg.attach(MIMEText(html_content, 'html'))
            
            if logo_path and os.path.exists(logo_path):
                with open(logo_path, 'rb') as img_file:
                    img = MIMEImage(img_file.read())
                    img.add_header('Content-ID', '<neoxam>')
                    img.add_header('Content-Disposition', 'inline', filename=os.path.basename(logo_path))
                    msg.attach(img)
                    print(f"[DEBUG] EmailManager: Attached logo {logo_path} to email")
            
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)
            
            print(f"[INFO] EmailManager: Acceptance email sent to {recipient_email}")
            return True, "Email de validation envoyé."
        except smtplib.SMTPAuthenticationError as e:
            print(f"[ERROR] EmailManager: SMTP authentication error: {e}")
            return False, "Erreur d'authentification SMTP. Vérifiez l'email et le mot de passe."
        except smtplib.SMTPException as e:
            print(f"[ERROR] EmailManager: SMTP error sending acceptance email: {e}")
            return False, f"Erreur SMTP: {str(e)}"
        except Exception as e:
            print(f"[ERROR] EmailManager: General error sending acceptance email: {e}")
            return False, f"Erreur lors de l'envoi de l'email: {str(e)}"


class EmailSendThread(QThread):
    """Thread pour envoyer des emails de manière asynchrone."""
    
    email_sent = pyqtSignal(bool, str)
    
    def __init__(self, email_manager, recipient_email, username, cover_image_path=None, email_type="confirmation"):
        """Initialise le thread d'envoi d'email."""
        super().__init__()
        self.email_manager = email_manager
        self.recipient_email = recipient_email
        self.username = username
        self.cover_image_path = cover_image_path
        self.email_type = email_type
        print(f"[DEBUG] EmailSendThread initialized for {email_type} to {recipient_email} with cover_image_path={cover_image_path}")
    
    def run(self):
        """Exécute l'envoi d'email dans le thread."""
        try:
            if self.email_type == "confirmation":
                success, message = self.email_manager.send_confirmation_email(
                    self.recipient_email, self.username, logo_path=self.cover_image_path
                )
            elif self.email_type == "acceptance":
                success, message = self.email_manager.send_acceptance_email(
                    self.recipient_email, self.username, logo_path=self.cover_image_path
                )
            else:
                success, message = False, "Type d'email non supporté"
            self.email_sent.emit(success, message)
            print(f"[DEBUG] EmailSendThread: Email sent signal emitted - success={success}, message={message}")
        except Exception as e:
            self.email_sent.emit(False, f"Erreur dans le thread d'envoi: {str(e)}")
            print(f"[ERROR] EmailSendThread: Error in thread: {e}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    db_manager = DatabaseManager()
    
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec_())